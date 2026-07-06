#!/usr/bin/env python3
"""
GloBE XML Export App
Streamlit web UI for converting the Swiss QDMTT Excel template to OECD GIR XML.
Run: streamlit run globe_xml_app.py
"""

import io
import json
import logging
import os
import re
import uuid
import zipfile
import openpyxl
import xml.etree.ElementTree as ET
from datetime import date, datetime, timezone

from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding as sym_padding
from cryptography.hazmat.primitives.asymmetric import padding as asym_padding
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives.serialization import load_pem_public_key
from cryptography.x509 import load_pem_x509_certificate

import base64
from PIL import Image
import streamlit as st

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)

VERSION = "2.7.1"   # correction mode: XML-upload dropzone gets its own headline (was showing the template text via the global ::after CSS)

# ─── XML SETUP ───────────────────────────────────────────────────────────────

GIR_NS = "urn:oecd:ties:globe:v2"        # confirmed from GLOBEXML_v1.0.xsd targetNamespace
STF_NS = "urn:oecd:ties:globestf:v5"    # DocSpec children (DocTypeIndic, DocRefId)
XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"
N = "{" + GIR_NS + "}"
S = "{" + STF_NS + "}"
ET.register_namespace("globe", GIR_NS)
ET.register_namespace("stf", STF_NS)
ET.register_namespace("xsi", XSI_NS)


# ─── MAPPINGS ────────────────────────────────────────────────────────────────

DATA_COL = "N"

INCOME_ADJUSTMENTS: dict[int, str] = {
    238: "GIR2001", 239: "GIR2002", 240: "GIR2003", 241: "GIR2004",
    242: "GIR2005", 243: "GIR2006", 244: "GIR2007", 245: "GIR2008",
    246: "GIR2009", 247: "GIR2010", 248: "GIR2011", 249: "GIR2012",
    250: "GIR2013", 251: "GIR2014", 252: "GIR2015", 253: "GIR2016",
    254: "GIR2017", 255: "GIR2018", 256: "GIR2019", 257: "GIR2020",
    258: "GIR2021", 259: "GIR2022", 260: "GIR2023", 261: "GIR2024",
    262: "GIR2025", 263: "GIR2026",
}

COVERED_TAX_ADJUSTMENTS: dict[int, str] = {
    297: "GIR2701", 298: "GIR2703", 299: "GIR2704", 300: "GIR2705",
    301: "GIR2706", 302: "GIR2707", 303: "GIR2708", 304: "GIR2709",
    305: "GIR2710", 306: "GIR2711", 307: "GIR2712", 308: "GIR2713",
    309: "GIR2714", 310: "GIR2715", 311: "GIR2716", 312: "GIR2717",
    313: "GIR2718",
}

GIR_INCOME_LABELS: dict[str, str] = {
    "GIR2001": "Net Taxes Expense (Art. 3.2.1a)",
    "GIR2002": "Excluded Dividends (Art. 3.2.1b)",
    "GIR2003": "Excluded Equity Gain or Loss (Art. 3.2.1c)",
    "GIR2004": "Included Revaluation Method Gain or Loss (Art. 3.2.1d)",
    "GIR2005": "Gain/Loss on Disposal of Assets excluded under Art. 6.3 (Art. 3.2.1e)",
    "GIR2006": "Asymmetric Foreign Currency Gains or Losses (Art. 3.2.1f)",
    "GIR2007": "Policy Disallowed Expenses (Art. 3.2.1g)",
    "GIR2008": "Prior Period Errors (Art. 3.2.1h)",
    "GIR2009": "Changes in Accounting Principles (Art. 3.2.1h)",
    "GIR2010": "Accrued Pension Expense (Art. 3.2.1i)",
    "GIR2011": "Debt Releases (Art. 3.2.1)",
    "GIR2012": "Stock-based Compensation (Art. 3.2.2)",
    "GIR2013": "Arm's Length Adjustments (Art. 3.2.3)",
    "GIR2014": "QRTC / Marketable Transferable Tax Credit (Art. 3.2.4)",
    "GIR2015": "Election for Gains/Losses – Realisation Principle (Art. 3.2.5)",
    "GIR2016": "Election for Adjusted Asset Gain (Art. 3.2.6)",
    "GIR2017": "Intragroup Financing Arrangement Expense (Art. 3.2.7)",
    "GIR2018": "Election for Intragroup Transactions – Same Jurisdiction (Art. 3.2.8)",
    "GIR2019": "Insurance Company Taxes Charged to Policyholders (Art. 3.2.9)",
    "GIR2020": "AT1/RT1 Capital Distribution Adjustments (Art. 3.2.10)",
    "GIR2021": "CE Joining/Leaving MNE Group (Art. 3.2.11 & 6.2)",
    "GIR2022": "Reduction of GloBE Income – UPE Flow-through Entity (Art. 3.2.11 & 7.1)",
    "GIR2023": "Reduction of GloBE Income – UPE Deductible Dividend Regime (Art. 3.2.11 & 7.2)",
    "GIR2024": "Taxable Distribution Method Election (Art. 3.2.11 & 7.6)",
    "GIR2025": "International Shipping Income (Art. 3.3)",
    "GIR2026": "Transactions between Constituent Entities (Art. 9.1.3)",
}

GIR_TAX_LABELS: dict[str, str] = {
    "GIR2701": "Covered Tax Accrued as Expense (Art. 4.1.2a)",
    "GIR2702": "GloBE Loss Deferred Tax Asset (Art. 4.1.2b & 4.5.3)",
    "GIR2703": "Covered Taxes – Uncertain Tax Position Prior Year (Art. 4.1.2c)",
    "GIR2704": "QRTC / Marketable Transferable Tax Credit – Current Tax Reduction (Art. 4.1.2d)",
    "GIR2705": "Qualified Flow-through Tax Benefits (Art. 3.2.1c)",
    "GIR2706": "Current Tax on Excluded Income (Art. 4.1.3a)",
    "GIR2707": "Non-QRTC / Other Tax Credits (Art. 4.1.3b)",
    "GIR2708": "Covered Taxes Refunded or Credited (Art. 4.1.3c)",
    "GIR2709": "Current Tax – Uncertain Tax Position (Art. 4.1.3d)",
    "GIR2710": "Current Tax Not Expected Paid within 3 Years (Art. 4.1.3e)",
    "GIR2711": "Post-filing Adjustments (Art. 4.6.1)",
    "GIR2712": "Covered Taxes – Net Asset Gain/Loss (Art. 3.2.6)",
    "GIR2713": "Reduction – UPE Flow-through Entity (Art. 7.1)",
    "GIR2714": "Covered Taxes – UPE Deductible Dividend Regime (Art. 7.2.2)",
    "GIR2715": "Deemed Distribution Tax (Art. 7.3)",
    "GIR2716": "Taxable Distribution Method Election (Art. 7.6b)",
    "GIR2717": "Total Deferred Tax Adjustment Amount (Art. 4.4.1b)",
    "GIR2718": "Covered Taxes in Equity / OCI (Art. 4.1.1c)",
    "GIR2719": "Excess Negative Tax Expense Carry Forward Generated (Art. 4.1.5 & 5.2.1)",
    "GIR2720": "Excess Negative Tax Expense Carry Forward Utilized (Art. 4.1.5 & 5.2.1)",
}

UPE_RULES_OPTIONS = ["GIR201", "GIR202", "GIR203", "GIR204", "GIR205"]
UPE_RULES_LABELS: dict[str, str] = {
    "GIR201": "GIR201 — QIIR (other jurisdictions only)",
    "GIR202": "GIR202 — QIIR (other + parent jurisdiction)",
    "GIR203": "GIR203 — QUTPR",
    "GIR204": "GIR204 — QDMTT",
    "GIR205": "GIR205 — Not applicable",
}

UPE_GLOBE_STATUS_OPTIONS = [
    "GIR301", "GIR302", "GIR303", "GIR304", "GIR305", "GIR306",
    "GIR307", "GIR308", "GIR309", "GIR310", "GIR311", "GIR312",
    "GIR313", "GIR314", "GIR315", "GIR316", "GIR317", "GIR318",
]
UPE_GLOBE_STATUS_LABELS: dict[str, str] = {
    "GIR301": "GIR301 — Constituent Entity",
    "GIR302": "GIR302 — Flow-Through Entity – Tax Transparent",
    "GIR303": "GIR303 — Flow-Through Entity – Reverse Hybrid",
    "GIR304": "GIR304 — Hybrid Entity",
    "GIR305": "GIR305 — Permanent Establishment",
    "GIR306": "GIR306 — Main Entity",
    "GIR307": "GIR307 — Minority-Owned Parent Entity",
    "GIR308": "GIR308 — Minority-Owned Subsidiary",
    "GIR309": "GIR309 — Minority-Owned Constituent Entity",
    "GIR310": "GIR310 — Investment Entity",
    "GIR311": "GIR311 — Insurance Investment Entity",
    "GIR312": "GIR312 — Securitisation Entity",
    "GIR313": "GIR313 — JV",
    "GIR314": "GIR314 — JV Subsidiary",
    "GIR315": "GIR315 — Non-Material Constituent Entity",
    "GIR316": "GIR316 — Excluded Entity",
    "GIR317": "GIR317 — Parent Entity (Art. 10.3.5)",
    "GIR318": "GIR318 — Non-group Member",
}

ROW_EXCESS_NEG_GENERATED = 95
ROW_EXCESS_NEG_UTILIZED  = 96
EXCESS_NEG_COL = "H"
ROW_ADJUSTED_FANIL     = 236
ROW_NET_GLOBE_INCOME   = 264
ROW_AGGREGATE_CURR_TAX = 295
ROW_ADJUSTED_COV_TAX   = 314


# ─── UI OPTIONS ──────────────────────────────────────────────────────────────

CURRENCIES = [
    "CHF", "EUR", "USD", "GBP", "JPY", "AUD", "CAD", "SEK", "NOK", "DKK",
    "SGD", "HKD", "NZD", "CNY", "INR", "BRL", "MXN", "KRW", "ZAR", "RUB",
    "PLN", "CZK", "HUF", "RON", "BGN", "ISK", "TRY", "SAR", "AED",
    "ILS", "THB", "IDR", "MYR", "PHP", "CLP", "ARS", "COP", "PEN",
    "EGP", "NGN", "KES", "MAD",
]

FAS_OPTIONS = [
    "Swiss GAAP FER",
    "IFRS",
    "US GAAP",
    "UK GAAP",
    "HGB",
    "Local GAAP",
]

_COUNTRIES = [
    ("AD", "Andorra"), ("AE", "United Arab Emirates"), ("AL", "Albania"),
    ("AM", "Armenia"), ("AO", "Angola"), ("AR", "Argentina"), ("AT", "Austria"),
    ("AU", "Australia"), ("AZ", "Azerbaijan"), ("BA", "Bosnia and Herzegovina"),
    ("BB", "Barbados"), ("BE", "Belgium"), ("BG", "Bulgaria"), ("BH", "Bahrain"),
    ("BM", "Bermuda"), ("BR", "Brazil"), ("BS", "Bahamas"), ("BW", "Botswana"),
    ("BY", "Belarus"), ("BZ", "Belize"), ("CA", "Canada"), ("CH", "Switzerland"),
    ("CL", "Chile"), ("CN", "China"), ("CO", "Colombia"), ("CR", "Costa Rica"),
    ("CY", "Cyprus"), ("CZ", "Czech Republic"), ("DE", "Germany"), ("DK", "Denmark"),
    ("DZ", "Algeria"), ("EE", "Estonia"), ("EG", "Egypt"), ("ES", "Spain"),
    ("FI", "Finland"), ("FR", "France"), ("GB", "United Kingdom"), ("GE", "Georgia"),
    ("GH", "Ghana"), ("GI", "Gibraltar"), ("GR", "Greece"), ("GT", "Guatemala"),
    ("HK", "Hong Kong"), ("HR", "Croatia"), ("HU", "Hungary"), ("ID", "Indonesia"),
    ("IE", "Ireland"), ("IL", "Israel"), ("IN", "India"), ("IS", "Iceland"),
    ("IT", "Italy"), ("JE", "Jersey"), ("JM", "Jamaica"), ("JO", "Jordan"),
    ("JP", "Japan"), ("KE", "Kenya"), ("KG", "Kyrgyzstan"), ("KR", "South Korea"),
    ("KW", "Kuwait"), ("KZ", "Kazakhstan"), ("LB", "Lebanon"), ("LI", "Liechtenstein"),
    ("LT", "Lithuania"), ("LU", "Luxembourg"), ("LV", "Latvia"), ("MA", "Morocco"),
    ("MC", "Monaco"), ("MD", "Moldova"), ("ME", "Montenegro"), ("MK", "North Macedonia"),
    ("MT", "Malta"), ("MU", "Mauritius"), ("MX", "Mexico"), ("MY", "Malaysia"),
    ("NA", "Namibia"), ("NG", "Nigeria"), ("NL", "Netherlands"), ("NO", "Norway"),
    ("NZ", "New Zealand"), ("OM", "Oman"), ("PA", "Panama"), ("PE", "Peru"),
    ("PH", "Philippines"), ("PK", "Pakistan"), ("PL", "Poland"), ("PT", "Portugal"),
    ("QA", "Qatar"), ("RO", "Romania"), ("RS", "Serbia"), ("RU", "Russia"),
    ("SA", "Saudi Arabia"), ("SE", "Sweden"), ("SG", "Singapore"), ("SI", "Slovenia"),
    ("SK", "Slovakia"), ("SM", "San Marino"), ("TH", "Thailand"), ("TN", "Tunisia"),
    ("TR", "Turkey"), ("TT", "Trinidad and Tobago"), ("TW", "Taiwan"),
    ("UA", "Ukraine"), ("UG", "Uganda"), ("US", "United States"), ("UY", "Uruguay"),
    ("UZ", "Uzbekistan"), ("VN", "Vietnam"), ("ZA", "South Africa"),
    ("ZM", "Zambia"), ("ZW", "Zimbabwe"),
]
COUNTRY_DISPLAY = [f"{code} – {name}" for code, name in _COUNTRIES]


def _country_idx(code: str) -> int:
    for i, (c, _) in enumerate(_COUNTRIES):
        if c == code:
            return i
    return 0


# ─── HELPERS ─────────────────────────────────────────────────────────────────

def cell_int(ws, row: int, col: str = DATA_COL) -> int:
    v = ws[f"{col}{row}"].value
    if v is None:
        return 0
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return 0


def fmt_etr(adj_covered_tax: int, net_globe_income: int) -> str:
    if not net_globe_income:
        return "0.0000"
    rate = max(0.0, min(1.0, adj_covered_tax / net_globe_income))
    return f"{rate:.4f}"


def sub(parent: ET.Element, tag: str, text=None, **attrib) -> ET.Element:
    el = ET.SubElement(parent, N + tag, attrib)
    if text is not None:
        el.text = str(text)
    return el


def tin_element(parent: ET.Element, tin_value, issued_by=None, tin_type="GIR3001") -> ET.Element:
    """Emit a globe:TIN element per ESTV business rules.
    - Known TIN: value + TypeOfTIN (GIR3001) + issuedBy (both required, errors 70005/70004).
    - Unknown TIN: value 'NOTIN', TypeOfTIN='GIR3004', unknown='true', NO issuedBy
      (errors 70002/70003). NB: ESTV forbids an unknown TIN on a constituent entity's
      own ID (error 70006) — that is enforced as a structural-validation failure, not here."""
    if tin_value:
        attrib = {"TypeOfTIN": tin_type or "GIR3001"}
        if issued_by:
            attrib["issuedBy"] = issued_by
        el = ET.SubElement(parent, N + "TIN", attrib)
        el.text = str(tin_value)
    else:
        el = ET.SubElement(parent, N + "TIN", {"TypeOfTIN": "GIR3004", "unknown": "true"})
        el.text = "NOTIN"
    return el


# ─── CORE CONVERSION ─────────────────────────────────────────────────────────

def read_excel(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["QDMTT 2024"]

    data = {
        "adjusted_fanil":     cell_int(ws, ROW_ADJUSTED_FANIL),
        "net_globe_income":   cell_int(ws, ROW_NET_GLOBE_INCOME),
        "aggregate_curr_tax": cell_int(ws, ROW_AGGREGATE_CURR_TAX),
        "adjusted_cov_tax":   cell_int(ws, ROW_ADJUSTED_COV_TAX),
        "income_adj":         [],
        "cov_tax_adj":        [],
    }

    for row, gir_code in INCOME_ADJUSTMENTS.items():
        data["income_adj"].append((gir_code, cell_int(ws, row)))

    for row, gir_code in COVERED_TAX_ADJUSTMENTS.items():
        data["cov_tax_adj"].append((gir_code, cell_int(ws, row)))

    gen  = cell_int(ws, ROW_EXCESS_NEG_GENERATED, EXCESS_NEG_COL)
    util = cell_int(ws, ROW_EXCESS_NEG_UTILIZED,  EXCESS_NEG_COL)
    data["cov_tax_adj"].append(("GIR2719", gen))
    data["cov_tax_adj"].append(("GIR2720", util))

    return data


# ─── NEW TEMPLATE (2026 multi-sheet GIR) PARSING ─────────────────────────────
#
# The 2026 "GIR Template" workbook replaces the single "QDMTT 2024" sheet with:
#   "1 MNE Group Information"  — filing CE, MNE group, full corporate structure
#   "2 Safe Harbours XX" × N   — per-jurisdiction safe-harbour election (Phase 2)
#   "3 GloBE Computations"     — the computed jurisdiction's ETR (col E)
#
# Sheet 3 embeds XML element names next to each value (cols F/G), but at least one
# of those embedded codes is wrong (row 91 reads GIR2417 where the deferred-tax
# adjustment must be GIR2717). The line ORDER, however, is the fixed official OECD
# layout, so we map by verified row → code (authoritative) and only use the
# embedded labels as a soft sanity check. This mirrors the proven v1 approach.

COMP_SHEET  = "3 GloBE Computations"
MNE_SHEET   = "1 MNE Group Information"
COMP_COL    = "E"                      # values column on sheet 3

NEW_ROW_ADJUSTED_FANIL   = 40          # "Aggregate FANIL amount after allocations"
NEW_ROW_NET_GLOBE_INCOME = 68          # "Net GloBE Income (Loss) of the Jurisdiction"
NEW_ROW_AGG_CURRENT_TAX  = 73          # "Aggregate Current tax expense ..."
NEW_ROW_ADJUSTED_COV_TAX = 95          # "Adjusted Covered Taxes"
NEW_JUR_NAME_CELL        = "E10"       # "1. Name of the jurisdiction"

# rows 42..67  → GIR2001..GIR2026 (income adjustments, sheet-3 section 3.2.1.1)
NEW_INCOME_ADJUSTMENTS = {42 + i: f"GIR{2001 + i}" for i in range(26)}
# rows 75..94  → GIR2701..GIR2720 (covered-tax adjustments, section 3.2.1.2);
# fully sequential here, which neatly avoids the row-91 label anomaly.
NEW_COVERED_TAX_ADJUSTMENTS = {75 + i: f"GIR{2701 + i}" for i in range(20)}

# Reverse name → ISO-3166 alpha-2 map, derived from the existing _COUNTRIES list,
# plus a few aliases the official template may use.
_NAME_TO_ISO = {name.lower(): code for code, name in _COUNTRIES}
_NAME_TO_ISO.update({
    "usa": "US", "united states of america": "US",
    "uk": "GB", "great britain": "GB",
    "south korea": "KR", "korea": "KR",
    "russia": "RU", "russian federation": "RU",
    "czechia": "CZ",
})


def _iso_from_name(name) -> str | None:
    if not name:
        return None
    s = str(name).strip()
    # ISO 3166-1 alpha-2 (two letters) OR an OECD GloBE special code such as
    # X5 = Stateless (schema isoglobetypes allows X-series). Without the digit
    # branch, "X5" failed the match, returned None, and stateless PEs silently
    # fell back to the filing jurisdiction (causing an ESTV 70012 Rules mismatch).
    if re.fullmatch(r"[A-Z][A-Z0-9]", s):
        return s
    return _NAME_TO_ISO.get(s.lower())


def _gir_code(value) -> str | None:
    """Extract a leading GIRxxxx code from a free-text cell ('GIR202 – QIIR …')."""
    if value is None:
        return None
    m = re.search(r"GIR\d{3,4}", str(value))
    return m.group(0) if m else None


def _clean_tin(value) -> str | None:
    """Template uses 0 / blank / 'not applicable' for 'no TIN'. Return clean TIN or None."""
    if value in (None, 0, "0", ""):
        return None
    s = str(value).strip()
    if not s or s.lower() in ("not applicable", "n/a", "na", "none"):
        return None
    return s


_MONTHS = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
    "december": 12,
    # German fallback
    "januar": 1, "februar": 2, "märz": 3, "mai": 5, "juni": 6, "juli": 7,
    "oktober": 10, "dezember": 12,
}


def _parse_template_date(value) -> str | None:
    """Parse '1. January 2024' / '31. December 2024' (or an Excel date) → 'YYYY-MM-DD'."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    m = re.match(r"(\d{1,2})\.?\s+([A-Za-zäöü]+)\.?\s+(\d{4})", s)
    if m:
        day, month_name, year = m.groups()
        month = _MONTHS.get(month_name.lower())
        if month:
            return f"{int(year):04d}-{month:02d}-{int(day):02d}"
    return None


def read_mne_info(wb) -> dict:
    """Read filing/company/UPE details + the constituent-entity list from sheet 1."""
    ws = wb[MNE_SHEET]

    def g(addr):
        return ws[addr].value

    upe_jur_iso = _iso_from_name(g("F6")) or _iso_from_name(g("C24")) or "CH"

    mne = {
        "company_name":   (g("C26") or g("C6") or "").strip(),     # UPE / Filing CE name
        "tin_value":      (_clean_tin(g("C27")) or _clean_tin(g("D6")) or ""),
        "tin_issued_by":  upe_jur_iso,
        "jurisdiction":   upe_jur_iso,
        "filing_role":    _gir_code(g("E6")) or "GIR401",
        "group_name":     (g("B13") or "").strip(),
        "period_start":   _parse_template_date(g("C13")),
        "period_end":     _parse_template_date(g("D13")),
        "fas":            (str(g("C18")).strip() if g("C18") else None),
        "currency":       (str(g("D18")).strip().upper() if g("D18") else None),
        "upe_rules":      _gir_code(g("C25")) or "GIR204",
        "upe_globe_status": _gir_code(g("C29")) or "GIR301",
        "constituent_entities": [],
    }

    # 1.3.2.1 Constituent Entities (other than the UPE): columns C.. across,
    # fixed rows — jurisdiction r37, rules r38, name r40, TIN r41, status r43,
    # ownership type r45, ownership % r47. The UPE may also be listed here, so
    # exclude any row that matches the UPE (by TIN, else by name) to avoid
    # emitting it twice in the corporate structure.
    upe_tin_norm  = (mne["tin_value"] or "").replace(" ", "").lower()
    upe_name_norm = mne["company_name"].strip().lower()
    for col in range(3, ws.max_column + 1):
        name = ws.cell(row=40, column=col).value
        jur  = ws.cell(row=37, column=col).value
        if not name or not jur:
            continue
        tin = _clean_tin(ws.cell(row=41, column=col).value)
        is_upe = (
            (tin and tin.replace(" ", "").lower() == upe_tin_norm and upe_tin_norm)
            or str(name).strip().lower() == upe_name_norm
        )
        if is_upe:
            continue
        mne["constituent_entities"].append({
            "name":           str(name).strip(),
            "iso":            _iso_from_name(jur),
            "tin":            tin,
            "rules":          _gir_code(ws.cell(row=38, column=col).value) or "GIR205",
            "globe_status":   _gir_code(ws.cell(row=43, column=col).value) or "GIR301",
            "ownership_type": _gir_code(ws.cell(row=45, column=col).value) or "GIR802",
            "owner_tin":      _clean_tin(ws.cell(row=46, column=col).value),  # TIN of the holder
            "ownership_pct":  ws.cell(row=47, column=col).value,
        })

    return mne


def _computation_sheets(wb) -> list:
    """All sheet names holding a GloBE computation: the legacy single bare
    '3 GloBE Computations' tab and/or the per-jurisdiction suffixed
    '3 GloBE Computations XX' tabs (multi-jurisdiction templates)."""
    return [n for n in wb.sheetnames
            if n == COMP_SHEET or n.startswith(COMP_SHEET + " ")]


def _read_computation_ws(ws) -> dict:
    """Read one '3 GloBE Computations' worksheet (col E, row-mapped)."""
    data = {
        "jur_iso":            _iso_from_name(ws[NEW_JUR_NAME_CELL].value),
        "adjusted_fanil":     cell_int(ws, NEW_ROW_ADJUSTED_FANIL,   COMP_COL),
        "net_globe_income":   cell_int(ws, NEW_ROW_NET_GLOBE_INCOME, COMP_COL),
        "aggregate_curr_tax": cell_int(ws, NEW_ROW_AGG_CURRENT_TAX,  COMP_COL),
        "adjusted_cov_tax":   cell_int(ws, NEW_ROW_ADJUSTED_COV_TAX, COMP_COL),
        "income_adj":         [],
        "cov_tax_adj":        [],
    }

    for row, gir_code in NEW_INCOME_ADJUSTMENTS.items():
        data["income_adj"].append((gir_code, cell_int(ws, row, COMP_COL)))

    for row, gir_code in NEW_COVERED_TAX_ADJUSTMENTS.items():
        data["cov_tax_adj"].append((gir_code, cell_int(ws, row, COMP_COL)))

    return data


def read_computation(wb) -> dict:
    """Back-compat: read the first computation sheet as a single dict."""
    return _read_computation_ws(wb[_computation_sheets(wb)[0]])


def read_computations(wb) -> list:
    """Read every computation sheet into a list — one full-scope GloBE
    computation per jurisdiction. Supports both the single bare-named tab
    (legacy single-jurisdiction template) and the per-jurisdiction suffixed
    tabs. Tabs whose jurisdiction name (E10) is blank are unfilled and skipped."""
    out = []
    for name in _computation_sheets(wb):
        d = _read_computation_ws(wb[name])
        if not d["jur_iso"]:          # unfilled template tab → ignore
            continue
        out.append(d)
    return out


# ─── SAFE HARBOURS (sheet "2 Safe Harbours XX") ──────────────────────────────
SAFE_HARBOUR_PREFIX = "2 Safe Harbours"
SH_JUR_CELL      = "C6"    # name of the jurisdiction
SH_ELECTED_CELL  = "C17"   # "1. Safe Harbour elected" → GIR1201..GIR1209
SH_REVENUE_CELL  = "C37"   # Transitional CbCR: Total Revenue
SH_PROFIT_CELL   = "C38"   # Transitional CbCR: Profit (Loss) before Income Tax
SH_TAX_CELL      = "C39"   # Transitional CbCR: Simplified Covered Taxes
SH_CITRATE_CELL  = "C42"   # Transitional UTPR: Corporate income tax rate

# Which safe-harbour codes carry a Transitional CbCR Safe Harbour block.
CBCR_SH_CODES = {"GIR1203", "GIR1204", "GIR1205"}
UTPR_SH_CODES = {"GIR1206"}


def _cell_num(ws, addr):
    """Return an int for numeric cells, else None (blank/text)."""
    v = ws[addr].value
    if v is None or isinstance(v, str):
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def read_safe_harbours(wb) -> list:
    """Read every '2 Safe Harbours XX' tab that has a safe-harbour election.
    Tabs with no election (C17 empty) are skipped — partially-filled templates
    are normal, so empty jurisdiction tabs are simply ignored."""
    out = []
    for name in wb.sheetnames:
        if not name.startswith(SAFE_HARBOUR_PREFIX):
            continue
        ws = wb[name]
        sh_code = _gir_code(ws[SH_ELECTED_CELL].value)
        if not sh_code:                       # empty / unfilled tab → ignore
            continue
        jur_name = ws[SH_JUR_CELL].value
        cit = ws[SH_CITRATE_CELL].value
        try:
            cit = float(cit) if cit not in (None, "") else None
        except (TypeError, ValueError):
            cit = None
        out.append({
            "iso":        _iso_from_name(jur_name),
            "name":       str(jur_name).strip() if jur_name else "",
            "sh_code":    sh_code,
            "revenue":    _cell_num(ws, SH_REVENUE_CELL),
            "profit":     _cell_num(ws, SH_PROFIT_CELL),
            "income_tax": _cell_num(ws, SH_TAX_CELL),
            "cit_rate":   cit,
        })
    return out


def read_excel_v2(file_bytes: bytes) -> dict:
    """Parse the 2026 multi-sheet GIR template → {'mne', 'data', 'safe_harbours'}."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if MNE_SHEET not in wb.sheetnames:
        raise KeyError(MNE_SHEET)
    comps = read_computations(wb)
    if not comps:
        raise KeyError(COMP_SHEET)
    return {
        "mne":           read_mne_info(wb),
        "computations":  comps,        # one full-scope computation per jurisdiction
        "data":          comps[0],     # back-compat single-jurisdiction reference
        "safe_harbours": read_safe_harbours(wb),
    }


def build_xml(computations, cfg: dict, test_mode: bool = False) -> str:
    # Accept a single computation dict (legacy single-jurisdiction call) or a
    # list of computations (multi-jurisdiction template → one JurisdictionSection
    # with a full ETR computation per jurisdiction).
    if isinstance(computations, dict):
        computations = [computations]
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
    year = cfg["period_end"][:4]
    msg_ref = f"{cfg['jurisdiction']}{year}{cfg['jurisdiction']}{str(uuid.uuid4())}"

    root = ET.Element(N + "GLOBE_OECD", {
        "version": "1.0",
        "{" + XSI_NS + "}schemaLocation": "urn:oecd:ties:globe:v2 GLOBEXML_v1.0.xsd",
    })

    hdr = sub(root, "MessageSpec")
    sub(hdr, "SendingEntityIN",     cfg["tin_value"])
    sub(hdr, "TransmittingCountry", cfg["jurisdiction"])
    sub(hdr, "ReceivingCountry",    cfg["jurisdiction"])
    sub(hdr, "MessageType",         "GIR")
    sub(hdr, "MessageRefId",        msg_ref)
    sub(hdr, "MessageTypeIndic",    "GIR101")
    sub(hdr, "ReportingPeriod",     cfg["period_end"])
    sub(hdr, "Timestamp",           now)

    body = sub(root, "GLOBEBody")
    fi   = sub(body, "FilingInfo")

    filing_ce = sub(fi, "FilingCE")
    sub(filing_ce, "ResCountryCode", cfg["jurisdiction"])
    sub(filing_ce, "Name",           cfg["company_name"])
    sub(filing_ce, "TIN",            cfg["tin_value"],
                   issuedBy=cfg["tin_issued_by"], TypeOfTIN=cfg["tin_type"])
    sub(filing_ce, "Role",           cfg["reporting_role"])

    acct = sub(fi, "AccountingInfo")
    sub(acct, "CFSofUPE", cfg["cfs_of_upe"])
    sub(acct, "FAS",      cfg["fas"])
    sub(acct, "Currency", cfg["currency"])

    period = sub(fi, "Period")
    sub(period, "Start", cfg["period_start"])
    sub(period, "End",   cfg["period_end"])
    sub(fi, "NameMNE", cfg["company_name"])

    # OECD11 = new submission (Neumeldung) in test mode. OECD10 = resend (only for FilingInfo on corrections).
    doc_type_indic = "OECD11" if test_mode else "OECD1"
    doc_type_indic_sections = "OECD11" if test_mode else "OECD1"

    def add_docspec(parent):
        """Append a section DocSpec (DocTypeIndic + unique DocRefId)."""
        ds = sub(parent, "DocSpec")
        ET.SubElement(ds, S + "DocTypeIndic").text = doc_type_indic_sections
        ET.SubElement(ds, S + "DocRefId").text = f"{cfg['jurisdiction']}{year}-{str(uuid.uuid4())}"
        return ds

    fi_doc = sub(fi, "DocSpec")
    ET.SubElement(fi_doc, S + "DocTypeIndic").text = doc_type_indic
    ET.SubElement(fi_doc, S + "DocRefId").text = f"{cfg['jurisdiction']}{year}-{str(uuid.uuid4())}"

    gen_sec = sub(body, "GeneralSection")
    sub(gen_sec, "RecJurCode", cfg["rec_jur_code"])
    corp = sub(gen_sec, "CorporateStructure")
    upe_el = sub(corp, "UPE")
    other_upe = sub(upe_el, "OtherUPE")
    id_el = sub(other_upe, "ID")
    sub(id_el, "Name", cfg["company_name"])
    sub(id_el, "ResCountryCode", cfg["jurisdiction"])
    upe_tin = ET.SubElement(id_el, N + "TIN",
                            {"issuedBy": cfg["tin_issued_by"], "TypeOfTIN": cfg["tin_type"]})
    upe_tin.text = cfg["tin_value"]
    sub(id_el, "Rules", cfg["upe_rules"])
    sub(id_el, "GlobeStatus", cfg["upe_globe_status"])

    # Other constituent entities (CorporateStructure/CE*), one per non-UPE entity.
    # Map known CE TINs → jurisdiction so an owner reference can be issuedBy-matched.
    ce_iso_by_tin = {c["tin"]: c.get("iso") for c in cfg.get("constituent_entities", []) if c.get("tin")}
    for ce in cfg.get("constituent_entities", []):
        ce_el = sub(corp, "CE")
        ce_id = sub(ce_el, "ID")
        sub(ce_id, "Name", ce["name"])
        sub(ce_id, "ResCountryCode", ce.get("iso") or cfg["jurisdiction"])
        tin_element(ce_id, ce.get("tin"),
                    issued_by=ce.get("iso") or cfg["jurisdiction"],
                    tin_type=cfg["tin_type"])
        sub(ce_id, "Rules", ce.get("rules") or "GIR205")
        sub(ce_id, "GlobeStatus", ce.get("globe_status") or "GIR301")
        # Ownership (1..n, required). ESTV rule 70030: for GIR802/803/804 the owner
        # TIN must match a reported CE's ID/TIN. The template provides the owner TIN
        # in sheet-1 row 46; when present we use it (issuedBy matched to that CE).
        # When absent (no intermediate-ownership data), report UPE ownership (GIR801)
        # with the UPE's TIN, which ESTV accepts.
        own = sub(ce_el, "Ownership")
        otype     = ce.get("ownership_type") or "GIR802"
        owner_tin = ce.get("owner_tin")
        if otype in ("GIR802", "GIR803", "GIR804") and owner_tin:
            sub(own, "OwnershipType", otype)
            tin_element(own, owner_tin,
                        issued_by=ce_iso_by_tin.get(owner_tin) or ce.get("iso") or cfg["jurisdiction"],
                        tin_type=cfg["tin_type"])
        else:
            sub(own, "OwnershipType", "GIR801")
            tin_element(own, cfg["tin_value"],
                        issued_by=cfg["tin_issued_by"], tin_type=cfg["tin_type"])
        pct = ce.get("ownership_pct")
        try:
            pct_f = max(0.0, min(1.0, float(pct)))
        except (TypeError, ValueError):
            pct_f = 1.0
        sub(own, "OwnershipPercentage", f"{pct_f:.4f}")

    add_docspec(gen_sec)

    # ── Summaries — one per safe-harbour jurisdiction. Per GLOBEBody ordering, ALL
    #    Summary elements come after GeneralSection and before any JurisdictionSection.
    safe_harbours = cfg.get("safe_harbours", [])
    comp_isos        = [(c.get("jur_iso") or cfg["jurisdiction"]) for c in computations]
    comp_iso_set     = set(comp_isos)
    comp_by_iso      = {iso: c for iso, c in zip(comp_isos, computations)}
    comp_iso_default = comp_isos[0]
    # RecJurCode = the recipient/partner jurisdiction (CH for a domestic filing). Per
    # ESTV rule 98201 every section's RecJurCode must also appear in the GeneralSection,
    # which carries only this code — so all sections share it. The section's *own*
    # jurisdiction is reported via the Jurisdiction element, not RecJurCode.
    rec_jur = cfg["rec_jur_code"]
    for sh in safe_harbours:
        iso  = sh.get("iso") or comp_iso_default
        summ = sub(body, "Summary")
        sub(summ, "RecJurCode", rec_jur)
        sub(sub(summ, "Jurisdiction"), "JurisdictionName", iso)
        # ESTV rule 60024: a Summary carrying SafeHarbour/GLoBETut/… must also
        # provide JurWithTaxingRights/JurisdictionName (the jurisdiction itself).
        sub(sub(summ, "JurWithTaxingRights"), "JurisdictionName", iso)
        sub(summ, "SafeHarbour", sh["sh_code"])           # Summary order: SafeHarbour →
        if sh["sh_code"] == "GIR1202":                    # ETRRange → SBIE → QDMTTut → GLoBETut
            # ESTV rule 70043: a GIR1202 (QDMTT SH) Summary with JurWithTaxingRights
            # must also carry ETRRange, SBIE and QDMTTut.
            if iso in comp_by_iso and comp_by_iso[iso].get("net_globe_income"):
                _d = comp_by_iso[iso]
                _r = max(0.0, min(1.0, _d["adjusted_cov_tax"] / _d["net_globe_income"]))
                _band = f"GIR13{min(13, int(_r / 0.025) + 1):02d}"   # 2.5% bands; ≥30% → GIR1313
            else:
                _band = "GIR1314"                         # Section 3.2 not completed
            sub(summ, "ETRRange", _band)
            _sbie = sub(summ, "SBIE")
            sub(_sbie, "NotApplicable", "false")          # SBIE applies (CH has an SBIE amount)
            sub(_sbie, "NoTut", "true")                   # safe harbour ⇒ no top-up tax
            sub(summ, "QDMTTut", "GIR1401")               # QDMTT safe harbour: no QDMTT top-up tax
        sub(summ, "GLoBETut", "GIR1501")                  # safe harbour ⇒ GloBE top-up tax = 0
        add_docspec(summ)

    # ── Computed jurisdictions (sheet 3) — one full ETR computation per jurisdiction.
    #    A single-jurisdiction template yields one section; the multi-jurisdiction
    #    template (no safe harbours) yields one full JurisdictionSection per jurisdiction.
    for data in computations:
        comp_iso = data.get("jur_iso") or cfg["jurisdiction"]
        jur_sec = sub(body, "JurisdictionSection")
        sub(jur_sec, "RecJurCode",  rec_jur)
        sub(jur_sec, "Jurisdiction", comp_iso)

        globe_tax  = sub(jur_sec, "GLoBETax")
        etr        = sub(globe_tax, "ETR")
        etr_status = sub(etr, "ETRStatus")
        etr_comp   = sub(etr_status, "ETRComputation")
        oc         = sub(etr_comp, "OverallComputation")

        sub(oc, "FANIL",         data["adjusted_fanil"])
        sub(oc, "AdjustedFANIL", data["adjusted_fanil"])

        ngi = sub(oc, "NetGlobeIncome")
        sub(ngi, "Total", data["net_globe_income"])
        for gir_code, amount in data["income_adj"]:
            if amount != 0:
                adj = sub(ngi, "Adjustments")
                sub(adj, "Amount",         amount)
                sub(adj, "AdjustmentItem", gir_code)

        sub(oc, "IncomeTaxExpense",    data["aggregate_curr_tax"])
        sub(oc, "ETRRate",             fmt_etr(data["adjusted_cov_tax"], data["net_globe_income"]))
        sub(oc, "TopUpTaxPercentage",  "0.0000")

        act = sub(oc, "AdjustedCoveredTax")
        sub(act, "Total",                data["adjusted_cov_tax"])
        sub(act, "AggregrateCurrentTax", data["aggregate_curr_tax"])
        for gir_code, amount in data["cov_tax_adj"]:
            if amount != 0:
                adj = sub(act, "Adjustments")
                sub(adj, "Amount",         amount)
                sub(adj, "AdjustmentItem", gir_code)

        # ExcessProfits = NetGlobeIncome - SubstanceExclusion (SBIE=0; revisit per
        # jurisdiction once real SBIE figures are provided — see fix-later note).
        sub(oc, "ExcessProfits", data["net_globe_income"])

        qdmtt = sub(oc, "QDMTT")
        sub(qdmtt, "FAS",            cfg["fas"])
        sub(qdmtt, "Amount",         data["adjusted_cov_tax"])
        sub(qdmtt, "SBIEAvailable",  "false")
        sub(qdmtt, "DeMinAvailable", "false")
        sub(qdmtt, "Currency",       cfg["currency"])

        sub(oc, "TopUpTax", 0)

        ente = sub(oc, "ExcessNegTaxExpense")
        sub(ente, "PriorYearBalance", 0)
        sub(ente, "GeneratedInRFY",   0)
        sub(ente, "UtilizedInRFY",    0)
        sub(ente, "Remaining",        0)

        add_docspec(jur_sec)

    # ── Safe-harbour jurisdictions (every one except the computed jurisdiction,
    #    which already has its full computation section above).
    for sh in safe_harbours:
        iso = sh.get("iso") or comp_iso_default
        if iso in comp_iso_set:
            continue
        code   = sh["sh_code"]
        sh_sec = sub(body, "JurisdictionSection")
        sub(sh_sec, "RecJurCode",  rec_jur)
        sub(sh_sec, "Jurisdiction", iso)
        gt = sub(sh_sec, "GLoBETax")
        if code in CBCR_SH_CODES:
            etr = sub(gt, "ETR")
            # ESTV rule 70046: a Transitional CbCR Safe Harbour ETR must carry a
            # SubGroup with TypeofSubGroup GIR1607 (Constituent Entities). The
            # template has no real subgroup, so the filing entity's TIN identifies it.
            sg = sub(etr, "SubGroup")
            tin_element(sg, cfg["tin_value"], issued_by=cfg["tin_issued_by"], tin_type=cfg["tin_type"])
            sub(sg, "TypeofSubGroup", "GIR1607")
            ex   = sub(sub(etr, "ETRStatus"), "ETRException")
            cbcr = sub(ex, "TransitionalCbCRSafeHarbour")   # order: Revenue, Profit, IncomeTax
            if sh.get("revenue") is not None:
                sub(cbcr, "Revenue", sh["revenue"])
            sub(cbcr, "Profit", sh["profit"] if sh.get("profit") is not None else 0)
            if sh.get("income_tax") is not None:
                sub(cbcr, "IncomeTax", sh["income_tax"])
        elif code in UTPR_SH_CODES:
            etr = sub(gt, "ETR")
            sg  = sub(etr, "SubGroup")                       # GIR1609 = Transitional UTPR Safe Harbour
            tin_element(sg, cfg["tin_value"], issued_by=cfg["tin_issued_by"], tin_type=cfg["tin_type"])
            sub(sg, "TypeofSubGroup", "GIR1609")
            ex   = sub(sub(etr, "ETRStatus"), "ETRException")
            utpr = sub(ex, "UTPRSafeHarbour")
            rate = sh.get("cit_rate") or 0.0
            sub(utpr, "CITRate", f"{max(0.0, min(1.0, float(rate))):.4f}")
        # else GIR1201/1202/1207-1209: declaration lives in the Summary; GLoBETax stays empty.
        add_docspec(sh_sec)

    ET.indent(root, space="  ")
    raw = ET.tostring(root, encoding="unicode")
    return f"<?xml version='1.0' encoding='utf-8'?>\n{raw}"


# ─── ENCRYPTION ──────────────────────────────────────────────────────────────

def encrypt_for_estv(xml_str: str, pem_bytes: bytes) -> bytes:
    """
    Packages xml_str into the ESTV-required encrypted zip:
      Payload.xml → Payload.zip (compressed) → AES-256-CBC encrypted → "Payload"
      AES key + IV (48 bytes) → RSA PKCS#1 v1.5 encrypted → "Key"
      Final zip contains "Payload" + "Key"
    """
    # Step 1: compress XML into Payload.zip
    inner_buf = io.BytesIO()
    with zipfile.ZipFile(inner_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("Payload.xml", xml_str.encode("utf-8"))
    payload_zip = inner_buf.getvalue()

    # Step 2: AES-256 CBC encrypt the zip
    aes_key = os.urandom(32)
    iv      = os.urandom(16)
    padder  = sym_padding.PKCS7(128).padder()
    padded  = padder.update(payload_zip) + padder.finalize()
    cipher  = Cipher(algorithms.AES(aes_key), modes.CBC(iv), backend=default_backend())
    enc     = cipher.encryptor()
    encrypted_payload = enc.update(padded) + enc.finalize()

    # Step 3: RSA-encrypt the 48-byte key material (key || iv)
    try:
        pub_key = load_pem_public_key(pem_bytes, backend=default_backend())
    except Exception:
        cert    = load_pem_x509_certificate(pem_bytes, default_backend())
        pub_key = cert.public_key()
    encrypted_key = pub_key.encrypt(aes_key + iv, asym_padding.PKCS1v15())

    # Step 4: bundle into final zip
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_STORED) as zf:
        zf.writestr("Payload", encrypted_payload)
        zf.writestr("Key",     encrypted_key)
    return out_buf.getvalue()


# ─── VALIDATION ──────────────────────────────────────────────────────────────

def validate_xml(xml_str: str) -> list[tuple[str, bool, str]]:
    """Returns list of (label, passed, detail) tuples. File-validity only —
    submission mode (test/prod) is NOT a file-validation point and is surfaced
    separately as a badge on the Step 2 header + a banner at the mode toggle."""
    results = []

    def check(label, passed, detail=""):
        results.append((label, passed, detail))

    # 1. Well-formed XML
    try:
        root = ET.fromstring(xml_str)
        check("Well-formed XML", True)
    except ET.ParseError as e:
        check("Well-formed XML", False, str(e))
        return results

    def _nsp(path: str) -> str:
        return "/".join(N + s for s in path.split("/"))

    def text(path):
        el = root.find(_nsp(path))
        return el.text.strip() if el is not None and el.text else None

    def findall(path):
        return root.findall(_nsp(path))

    # 2. Root element
    check("Root element (GLOBE_OECD)",
          root.tag == N + "GLOBE_OECD" and root.find(N + "MessageSpec") is not None)

    # 3. MessageSpec — all required fields (incl. Swiss SendingEntityIN)
    hdr_fields = ["TransmittingCountry", "ReceivingCountry", "MessageType",
                  "MessageRefId", "MessageTypeIndic", "ReportingPeriod",
                  "Timestamp", "SendingEntityIN"]
    missing_hdr = [f for f in hdr_fields if text(f"MessageSpec/{f}") is None]
    check("MessageSpec — all required fields (incl. SendingEntityIN)", not missing_hdr,
          f"Missing: {', '.join(missing_hdr)}" if missing_hdr else "")

    # 4. MessageRefId format: CH[0-9]{4}CH...
    msg_ref = text("MessageSpec/MessageRefId")
    check("MessageRefId format (CH[year]CH[uuid])",
          bool(msg_ref and re.match(r"^[A-Z]{2}\d{4}[A-Z]{2}.+", msg_ref)),
          msg_ref or "missing")

    # 5. Timestamp format
    ts = text("MessageSpec/Timestamp")
    check("Timestamp format (YYYY-MM-DDTHH:MM:SS)",
          bool(ts and re.match(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}", ts)),
          ts or "missing")

    # 6. Period dates
    start = text("GLOBEBody/FilingInfo/Period/Start")
    end   = text("GLOBEBody/FilingInfo/Period/End")
    date_ok = bool(
        start and re.match(r"\d{4}-\d{2}-\d{2}$", start) and
        end   and re.match(r"\d{4}-\d{2}-\d{2}$", end)
    )
    check("Period dates (YYYY-MM-DD)", date_ok,
          f"Start: {start}  End: {end}" if not date_ok else "")

    # 7. Company name — not placeholder
    name = text("GLOBEBody/FilingInfo/FilingCE/Name")
    name_ok = bool(name and name != "PLACEHOLDER_COMPANY_AG")
    check("Company name (not placeholder)", name_ok,
          "Still set to PLACEHOLDER_COMPANY_AG" if not name_ok else "")

    # 8. Role in FilingCE (GIR401–GIR405)
    role = text("GLOBEBody/FilingInfo/FilingCE/Role")
    check("FilingCE Role (GIR401–GIR405)",
          bool(role and re.match(r"^GIR40[1-5]$", role)),
          role or "missing")

    # 9. TIN — not placeholder, has required attributes
    tin_el = root.find(_nsp("GLOBEBody/FilingInfo/FilingCE/TIN"))
    tin_val = tin_el.text.strip() if tin_el is not None and tin_el.text else None
    tin_ok = bool(tin_val and tin_val != "CHE-123456789")
    check("TIN (not placeholder)", tin_ok,
          "Still set to CHE-123456789" if not tin_ok else "")
    if tin_el is not None:
        check("TIN attributes (issuedBy + TypeOfTIN)",
              bool(tin_el.get("issuedBy") and tin_el.get("TypeOfTIN")))

    # 10. DocSpec in FilingInfo
    fi_doc = root.find(_nsp("GLOBEBody/FilingInfo/DocSpec"))
    fi_doc_ok = (
        fi_doc is not None and
        fi_doc.find(S + "DocTypeIndic") is not None and
        fi_doc.find(S + "DocRefId") is not None
    )
    check("FilingInfo DocSpec (DocTypeIndic + DocRefId)", fi_doc_ok)

    # 11. RecJurCode in JurisdictionSection
    rec_jur = text("GLOBEBody/JurisdictionSection/RecJurCode")
    check("JurisdictionSection RecJurCode present",
          bool(rec_jur and re.match(r"^[A-Z]{2}$", rec_jur)),
          rec_jur or "missing")

    # 12. DocSpec in JurisdictionSection
    jur_doc = root.find(_nsp("GLOBEBody/JurisdictionSection/DocSpec"))
    jur_doc_ok = (
        jur_doc is not None and
        jur_doc.find(S + "DocTypeIndic") is not None and
        jur_doc.find(S + "DocRefId") is not None
    )
    check("JurisdictionSection DocSpec (DocTypeIndic + DocRefId)", jur_doc_ok)

    # Currency — ISO 4217 text content
    ccy_el = root.find(_nsp("GLOBEBody/FilingInfo/AccountingInfo/Currency"))
    check("Currency (ISO 4217 text)",
          bool(ccy_el is not None and ccy_el.text and len(ccy_el.text.strip()) == 3))

    # OverallComputation — required elements
    oc = ("GLOBEBody/JurisdictionSection/GLoBETax/ETR"
          "/ETRStatus/ETRComputation/OverallComputation")
    oc_fields = ["FANIL", "AdjustedFANIL", "IncomeTaxExpense",
                 "ETRRate", "TopUpTaxPercentage"]
    missing_oc = [f for f in oc_fields if text(f"{oc}/{f}") is None]
    check("OverallComputation — required elements", not missing_oc,
          f"Missing: {', '.join(missing_oc)}" if missing_oc else "")

    # ETRRate — decimal 0–1, 4 decimal places
    etr_val = text(f"{oc}/ETRRate")
    try:
        etr_f  = float(etr_val) if etr_val else None
        etr_ok = (etr_f is not None and 0 <= etr_f <= 1
                  and bool(re.match(r"^\d\.\d{4}$", etr_val)))
    except (ValueError, TypeError):
        etr_ok = False
    check("ETRRate format (0.0000 – 1.0000)", etr_ok, etr_val or "missing")

    # TopUpTaxPercentage format
    tup = text(f"{oc}/TopUpTaxPercentage")
    check("TopUpTaxPercentage format (0.0000)",
          bool(tup and re.match(r"^\d\.\d{4}$", tup)), tup or "missing")

    # NetGlobeIncome — only valid GIR20xx codes (zero-value items are omitted)
    valid_ngi = {f"GIR{2000+i}" for i in range(1, 27)}
    ngi_codes = {el.text for el in findall(
        f"{oc}/NetGlobeIncome/Adjustments/AdjustmentItem") if el.text}
    invalid_ngi = ngi_codes - valid_ngi
    check("NetGlobeIncome — adjustment codes valid (GIR2001–GIR2026)", not invalid_ngi,
          f"Invalid: {', '.join(sorted(invalid_ngi))}" if invalid_ngi else "")

    # AdjustedCoveredTax — only valid GIR27xx codes
    valid_act = {f"GIR27{i:02d}" for i in range(1, 21)}
    act_codes = {el.text for el in findall(
        f"{oc}/AdjustedCoveredTax/Adjustments/AdjustmentItem") if el.text}
    invalid_act = act_codes - valid_act
    check("AdjustedCoveredTax — adjustment codes valid (GIR2701–GIR2720)", not invalid_act,
          f"Invalid: {', '.join(sorted(invalid_act))}" if invalid_act else "")

    # GeneralSection present with CorporateStructure
    gen_sec = root.find(_nsp("GLOBEBody/GeneralSection"))
    gen_ok = (
        gen_sec is not None and
        gen_sec.find(_nsp("CorporateStructure")) is not None and
        gen_sec.find(_nsp("CorporateStructure/UPE")) is not None
    )
    check("GeneralSection present (CorporateStructure / UPE)", gen_ok)

    # CorporateStructure — every CE carries a complete ID block (multi-jurisdiction)
    ce_ids = findall("GLOBEBody/GeneralSection/CorporateStructure/CE/ID")
    bad_ce = []
    for i, ce_id in enumerate(ce_ids, 1):
        req = ["Name", "ResCountryCode", "TIN", "Rules", "GlobeStatus"]
        missing = [r for r in req if ce_id.find(N + r) is None]
        if missing:
            name_el = ce_id.find(N + "Name")
            label = name_el.text if name_el is not None and name_el.text else f"CE#{i}"
            bad_ce.append(f"{label}: {', '.join(missing)}")
    check(f"CorporateStructure — CE entities complete (UPE + {len(ce_ids)} CE)",
          not bad_ce, "; ".join(bad_ce) if bad_ce else "")

    # Every constituent entity must carry a real identity TIN — ESTV rejects an
    # unknown TIN on a CE's own ID (error 70006). Hard requirement for this client.
    no_tin = []
    for i, ce_id in enumerate(ce_ids, 1):
        tin = ce_id.find(N + "TIN")
        if tin is not None and (tin.get("unknown") == "true"
                                or (tin.text or "").strip() == "NOTIN"):
            name_el = ce_id.find(N + "Name")
            no_tin.append(name_el.text if name_el is not None and name_el.text else f"CE#{i}")
    check("Every constituent entity has an identity TIN (ESTV 70006)",
          not no_tin,
          f"{len(no_tin)} without a TIN — complete in sheet 1: {', '.join(no_tin)}" if no_tin else "")

    # At least one JurisdictionSection carries a full ETR computation
    n_jur = len(findall("GLOBEBody/JurisdictionSection"))
    has_comp = root.find(_nsp(
        "GLOBEBody/JurisdictionSection/GLoBETax/ETR/ETRStatus/ETRComputation")) is not None
    check(f"JurisdictionSection(s) present ({n_jur}) with an ETR computation", has_comp)

    # Safe-harbour sections: each Summary must carry a SafeHarbour enum, and every
    # Summary jurisdiction should have a matching JurisdictionSection.
    summaries = findall("GLOBEBody/Summary")
    sh_codes  = [s.find(N + "SafeHarbour") for s in summaries]
    bad_sh    = [i + 1 for i, c in enumerate(sh_codes) if c is None or not (c.text or "").startswith("GIR12")]
    jur_codes = {j.text for j in findall("GLOBEBody/JurisdictionSection/Jurisdiction") if j.text}
    summ_jurs = {s.text for s in findall("GLOBEBody/Summary/Jurisdiction/JurisdictionName") if s.text}
    unmatched = sorted(summ_jurs - jur_codes)
    check(f"Safe-harbour Summaries ({len(summaries)}) valid & matched to a JurisdictionSection",
          not bad_sh and not unmatched,
          (f"bad SafeHarbour: {bad_sh}; " if bad_sh else "")
          + (f"Summary with no JurisdictionSection: {', '.join(unmatched)}" if unmatched else ""))

    # All amounts are integers
    non_int = [el.text for el in root.findall(".//" + N + "Amount")
               if el.text and "." in el.text]
    check("All amounts are integers (no decimals)", not non_int,
          f"Non-integer: {non_int[:3]}" if non_int else "")

    # ESTV 70012 — every entity resident in the same jurisdiction must carry the
    # SAME Rules value. A mixed jurisdiction (e.g. UPE GIR202 + CE GIR205) is the
    # exact rejection we hit; the file was only accepted once every CH entity
    # shared one Rules value. (Treated as strict-uniform — a false warning here is
    # far cheaper than a production rejection with no test portal to fall back on.)
    rules_by_jur: dict = {}
    corp_ids = (findall("GLOBEBody/GeneralSection/CorporateStructure/UPE/OtherUPE/ID")
                + findall("GLOBEBody/GeneralSection/CorporateStructure/UPE/ExcludedUPE/ID")
                + findall("GLOBEBody/GeneralSection/CorporateStructure/CE/ID"))
    for idel in corp_ids:
        res = idel.find(N + "ResCountryCode")
        rul = idel.find(N + "Rules")
        if res is not None and res.text and rul is not None and rul.text:
            rules_by_jur.setdefault(res.text, set()).add(rul.text)
    mixed = {j: sorted(v) for j, v in rules_by_jur.items() if len(v) > 1}
    check("Same-jurisdiction Rules consistency (ESTV 70012)",
          not mixed,
          "; ".join(f"{j}: {', '.join(v)}" for j, v in mixed.items()) if mixed else "")

    return results


# ─── CORRECTION MODE (Rektifikat) ────────────────────────────────────────────
# ESTV Technische Wegleitung Kap. 5.3.4–6.4: a correction message (GIR102) resends
# FilingInfo unchanged (OECD0 / test OECD10, SAME DocRefId — rule 60014) and replaces
# only the changed sections AS A WHOLE (OECD2 / test OECD12; Storno OECD3/OECD13),
# each with a fresh DocRefId + CorrDocRefId → the LAST accepted DocRefId (chains,
# rules 60005–60015). RecJurCode is locked to the original (98200); new and corrected
# sections must never be mixed in one message (60004). Unchanged sections are OMITTED.
#
# Design: corrections are a pure POST-PROCESSING step over a freshly built Neumeldung
# XML — build_xml()/validate_xml() are untouched, so parity with the accepted
# production filings is preserved by construction.

CORRECTABLE_TAGS = ("GeneralSection", "Summary", "JurisdictionSection", "UTPRAttribution")
_SECTION_ORDER   = {"FilingInfo": 0, "GeneralSection": 1, "Summary": 2,
                    "JurisdictionSection": 3, "UTPRAttribution": 4}
_TEST_INDICS     = {"OECD10", "OECD11", "OECD12", "OECD13"}


def _local_tag(tag: str) -> str:
    return tag.split("}", 1)[-1]


def _sec_jurisdiction(el) -> str:
    """Jurisdiction key of a section ('' for GeneralSection). Summary nests
    Jurisdiction/JurisdictionName; JurisdictionSection carries Jurisdiction text."""
    j = el.find(N + "Jurisdiction")
    if j is not None:
        if (j.text or "").strip():
            return j.text.strip()
        n = j.find(N + "JurisdictionName")
        if n is not None and (n.text or "").strip():
            return n.text.strip()
    return ""


def _walk_sections(body):
    """Yield (key, element) for every correctable section, keyed by
    type|jurisdiction (with #n suffix on duplicates), in document order."""
    seen: dict = {}
    for el in body:
        t = _local_tag(el.tag)
        if t not in CORRECTABLE_TAGS:
            continue
        base = f"{t}|{_sec_jurisdiction(el)}"
        n = seen.get(base, 0) + 1
        seen[base] = n
        yield (base if n == 1 else f"{base}#{n}"), el


def parse_original_xml(xml_bytes: bytes) -> dict:
    """Extract the correction context (DocRefIds per section) from the originally
    submitted RAW GIR XML. The _encrypted.zip cannot be used — it is encrypted with
    ESTV's public key. If the original was already corrected once, upload the LATEST
    accepted message instead (correction chains, Wegleitung 6.3.3)."""
    root = ET.fromstring(xml_bytes)
    if _local_tag(root.tag) != "GLOBE_OECD":
        raise ValueError("Not a GIR file — root element is not GLOBE_OECD")
    hdr  = root.find(N + "MessageSpec")
    body = root.find(N + "GLOBEBody")
    if hdr is None or body is None:
        raise ValueError("MessageSpec / GLOBEBody missing")

    def _h(tag):
        el = hdr.find(N + tag)
        return el.text.strip() if el is not None and el.text else None

    fi = body.find(N + "FilingInfo")
    if fi is None:
        raise ValueError("FilingInfo missing")
    fi_ds = fi.find(N + "DocSpec")
    if fi_ds is None or fi_ds.find(S + "DocRefId") is None:
        raise ValueError("FilingInfo DocSpec/DocRefId missing")

    indics = {(fi_ds.find(S + "DocTypeIndic").text or "").strip()
              if fi_ds.find(S + "DocTypeIndic") is not None else ""}
    ctx = {
        "message_ref_id":    _h("MessageRefId"),
        "message_type":      _h("MessageTypeIndic"),
        "reporting_period":  _h("ReportingPeriod"),
        "transmitting":      _h("TransmittingCountry") or "CH",
        "filing_doc_ref_id": (fi_ds.find(S + "DocRefId").text or "").strip(),
        "sections":          [],
        "root":              root,
    }
    for key, el in _walk_sections(body):
        ds = el.find(N + "DocSpec")
        if ds is None or ds.find(S + "DocRefId") is None or not (ds.find(S + "DocRefId").text or "").strip():
            raise ValueError(f"Section {key}: DocSpec/DocRefId missing")
        dti = ds.find(S + "DocTypeIndic")
        rec = el.find(N + "RecJurCode")
        t, jur = key.split("|")[0], key.split("|")[1].split("#")[0]
        ctx["sections"].append({
            "key":            key,
            "type":           t,
            "jurisdiction":   jur,
            "doc_ref_id":     ds.find(S + "DocRefId").text.strip(),
            "doc_type_indic": (dti.text or "").strip() if dti is not None else "",
            "rec_jur_code":   rec.text.strip() if rec is not None and rec.text else None,
            "element":        el,
        })
        indics.add(ctx["sections"][-1]["doc_type_indic"])
    ctx["test_mode"] = bool(indics & _TEST_INDICS)
    return ctx


def _canonical_section(el) -> str:
    """Content fingerprint of a section: DocSpec dropped, C14N with whitespace
    stripped — so UUID/indentation differences never count as a change."""
    clone = ET.fromstring(ET.tostring(el))
    for ds in clone.findall(N + "DocSpec"):
        clone.remove(ds)
    return ET.canonicalize(ET.tostring(clone, encoding="unicode"), strip_text=True)


def diff_correction_sections(original_ctx: dict, new_xml_str: str) -> list[dict]:
    """Per-section comparison of a freshly regenerated report vs the original.
    Status: 'changed' | 'unchanged' | 'new' (NOT allowed in a correction —
    needs a separate GIR101 Neumeldung, Wegleitung 6.4.2) | 'missing'
    (only in the original → Storno candidate). Advisory — the user decides."""
    body = ET.fromstring(new_xml_str).find(N + "GLOBEBody")
    new_by_key = dict(_walk_sections(body))
    rows = []
    for sec in original_ctx["sections"]:
        el = new_by_key.pop(sec["key"], None)
        if el is None:
            status = "missing"
        elif _canonical_section(el) == _canonical_section(sec["element"]):
            status = "unchanged"
        else:
            status = "changed"
        rows.append({"key": sec["key"], "type": sec["type"],
                     "jurisdiction": sec["jurisdiction"],
                     "doc_ref_id": sec["doc_ref_id"], "status": status})
    for key, el in new_by_key.items():
        t = key.split("|")[0]
        rows.append({"key": key, "type": t,
                     "jurisdiction": key.split("|")[1].split("#")[0],
                     "doc_ref_id": None, "status": "new"})
    return rows


def _insert_section(body, el) -> None:
    """Insert a section at its schema-valid position (GLOBEBody sequence:
    FilingInfo, GeneralSection, Summary*, JurisdictionSection*, UTPRAttribution*)."""
    rank = _SECTION_ORDER.get(_local_tag(el.tag), 99)
    for i, child in enumerate(body):
        if _SECTION_ORDER.get(_local_tag(child.tag), 99) > rank:
            body.insert(i, el)
            return
    body.append(el)


def apply_correction(new_xml_str: str, original_ctx: dict, selection: dict,
                     test_mode: bool = False) -> str:
    """Turn a freshly built Neumeldung XML into a GIR102 correction message.

    selection: {section key: 'correct' | 'delete'}. Sections not in the selection
    are dropped (unchanged sections are omitted from a correction). 'delete' for a
    section absent from the new XML resends the ORIGINAL element as the Storno body
    (Wegleitung 6.4.3: fill the Musselemente by resending the original content)."""
    root = ET.fromstring(new_xml_str)
    hdr  = root.find(N + "MessageSpec")
    body = root.find(N + "GLOBEBody")
    hdr.find(N + "MessageTypeIndic").text = "GIR102"

    year = (original_ctx.get("reporting_period") or "0000")[:4]
    cc   = original_ctx.get("transmitting") or "CH"
    dti_resend = "OECD10" if test_mode else "OECD0"
    dti_corr   = "OECD12" if test_mode else "OECD2"
    dti_del    = "OECD13" if test_mode else "OECD3"
    orig_by_key = {s["key"]: s for s in original_ctx["sections"]}

    def _new_docref() -> str:
        return f"{cc}{year}-{uuid.uuid4()}"

    def _rewrite_docspec(el, dti, corr_ref) -> None:
        ds = el.find(N + "DocSpec")
        ds.find(S + "DocTypeIndic").text = dti
        ds.find(S + "DocRefId").text = _new_docref()
        # stf DocSpec_Type sequence: DocTypeIndic, DocRefId, [CorrMessageRefId],
        # [CorrDocRefId] — appending CorrDocRefId last keeps the order valid.
        ET.SubElement(ds, S + "CorrDocRefId").text = corr_ref

    # FilingInfo → Resend with the ORIGINAL DocRefId, no CorrDocRefId (rule 60014/60012)
    fi_ds = body.find(N + "FilingInfo").find(N + "DocSpec")
    fi_ds.find(S + "DocTypeIndic").text = dti_resend
    fi_ds.find(S + "DocRefId").text = original_ctx["filing_doc_ref_id"]

    handled = set()
    for key, el in list(_walk_sections(body)):
        action = selection.get(key)
        orig   = orig_by_key.get(key)
        if action not in ("correct", "delete") or orig is None:
            body.remove(el)
            continue
        handled.add(key)
        rec = el.find(N + "RecJurCode")
        if rec is not None and orig.get("rec_jur_code"):
            rec.text = orig["rec_jur_code"]          # RecJurCode lock (98200)
        _rewrite_docspec(el, dti_del if action == "delete" else dti_corr,
                         orig["doc_ref_id"])

    # Storno of sections that only exist in the original: resend the original
    # element as a whole with a fresh DocSpec (OECD3/OECD13 + CorrDocRefId).
    for key, action in selection.items():
        if action != "delete" or key in handled or key not in orig_by_key:
            continue
        orig  = orig_by_key[key]
        clone = ET.fromstring(ET.tostring(orig["element"]))
        old_ds = clone.find(N + "DocSpec")
        if old_ds is not None:
            clone.remove(old_ds)
        ds = ET.SubElement(clone, N + "DocSpec")
        ET.SubElement(ds, S + "DocTypeIndic").text = dti_del
        ET.SubElement(ds, S + "DocRefId").text = _new_docref()
        ET.SubElement(ds, S + "CorrDocRefId").text = orig["doc_ref_id"]
        _insert_section(body, clone)

    ET.indent(root, space="  ")
    raw = ET.tostring(root, encoding="unicode")
    return f"<?xml version='1.0' encoding='utf-8'?>\n{raw}"


def validate_correction_xml(xml_str: str, original_ctx: dict) -> list[tuple[str, bool, str]]:
    """Layer 1 — the correction file itself. Every documented ESTV correction rule
    becomes a gating check (labels cite the rule)."""
    results = []

    def check(label, passed, detail=""):
        results.append((label, passed, detail))

    try:
        root = ET.fromstring(xml_str)
        check("Well-formed XML", True)
    except ET.ParseError as e:
        check("Well-formed XML", False, str(e))
        return results

    hdr  = root.find(N + "MessageSpec")
    body = root.find(N + "GLOBEBody")

    def _h(tag):
        el = hdr.find(N + tag) if hdr is not None else None
        return el.text.strip() if el is not None and el.text else None

    mti = _h("MessageTypeIndic")
    check("MessageTypeIndic = GIR102 (Korrekturmeldung)", mti == "GIR102", mti or "missing")

    mr = _h("MessageRefId")
    check("MessageRefId — new & valid format, ≠ original (rule 60001/6.3.2)",
          bool(mr and re.match(r"^[A-Z]{2}\d{4}[A-Z]{2}.+", mr)
               and mr != original_ctx.get("message_ref_id")),
          mr or "missing")

    rp = _h("ReportingPeriod")
    check("ReportingPeriod matches the original filing",
          rp == original_ctx.get("reporting_period"),
          f"{rp} vs original {original_ctx.get('reporting_period')}" if rp != original_ctx.get("reporting_period") else "")

    year = (original_ctx.get("reporting_period") or "0000")[:4]
    orig_docrefs   = {s["doc_ref_id"] for s in original_ctx["sections"]}
    orig_docrefs.add(original_ctx["filing_doc_ref_id"])
    orig_by_docref = {s["doc_ref_id"]: s for s in original_ctx["sections"]}

    # FilingInfo: OECD0/OECD10 resend, SAME DocRefId, no CorrDocRefId
    fi = body.find(N + "FilingInfo") if body is not None else None
    fi_ds  = fi.find(N + "DocSpec") if fi is not None else None
    fi_dti = fi_ds.find(S + "DocTypeIndic").text if fi_ds is not None and fi_ds.find(S + "DocTypeIndic") is not None else None
    fi_dr  = fi_ds.find(S + "DocRefId").text if fi_ds is not None and fi_ds.find(S + "DocRefId") is not None else None
    check("FilingInfo resent as OECD0/OECD10 with the ORIGINAL DocRefId (rule 60014)",
          fi_dti in ("OECD0", "OECD10") and fi_dr == original_ctx["filing_doc_ref_id"],
          f"DocTypeIndic={fi_dti}, DocRefId={fi_dr}")
    check("FilingInfo resend carries no CorrDocRefId (rule 60012)",
          fi_ds is not None and fi_ds.find(S + "CorrDocRefId") is None)

    # Sections
    secs = []
    for key, el in _walk_sections(body if body is not None else []):
        ds = el.find(N + "DocSpec")
        secs.append({
            "key":  key,
            "type": key.split("|")[0],
            "dti":  (ds.find(S + "DocTypeIndic").text or "").strip() if ds is not None and ds.find(S + "DocTypeIndic") is not None else "",
            "dr":   (ds.find(S + "DocRefId").text or "").strip() if ds is not None and ds.find(S + "DocRefId") is not None else "",
            "corr": (ds.find(S + "CorrDocRefId").text or "").strip() if ds is not None and ds.find(S + "CorrDocRefId") is not None else None,
            "rec":  el.find(N + "RecJurCode").text.strip() if el.find(N + "RecJurCode") is not None and el.find(N + "RecJurCode").text else None,
        })

    check("At least one corrected/deleted section present", len(secs) > 0)

    bad_mix = [s["key"] for s in secs if s["dti"] not in ("OECD2", "OECD3", "OECD12", "OECD13")]
    check("No new sections mixed in — all DocTypeIndic OECD2/OECD3 (rule 60004)",
          not bad_mix, ", ".join(bad_mix))

    all_indics = {s["dti"] for s in secs} | ({fi_dti} if fi_dti else set())
    flavors = {i in _TEST_INDICS for i in all_indics}
    check("Test/production DocTypeIndic flavor consistent (rule 50009)",
          len(flavors) <= 1, f"mixed: {sorted(all_indics)}" if len(flavors) > 1 else "")

    no_corr = [s["key"] for s in secs if not s["corr"]]
    check("Every section carries a CorrDocRefId (rule 60015)", not no_corr, ", ".join(no_corr))

    bad_target = []
    for s in secs:
        orig = orig_by_docref.get(s["corr"] or "")
        if orig is None or orig["type"] != s["type"]:
            bad_target.append(s["key"])
    check("Every CorrDocRefId resolves to an original section of the same type (rules 60005/60008)",
          not bad_target, ", ".join(bad_target))

    corr_ids = [s["corr"] for s in secs if s["corr"]]
    dup_corr = sorted({c for c in corr_ids if corr_ids.count(c) > 1})
    check("Each CorrDocRefId used only once in this message (rule 60006)",
          not dup_corr, ", ".join(dup_corr))

    new_ids = [s["dr"] for s in secs]
    reused  = sorted({d for d in new_ids if d in orig_docrefs or new_ids.count(d) > 1})
    check("New DocRefIds unique & never reused from the original (rule 60007)",
          not reused, ", ".join(reused))

    bad_fmt = [s["key"] for s in secs
               if not re.match(r"^[A-Z]{2}\d{4}.{1,194}$", s["dr"] or "")
               or (s["dr"] or "")[2:6] != year]
    check(f"DocRefId format CH{year}… — period matches the filing (rule 60011)",
          not bad_fmt, ", ".join(bad_fmt))

    bad_rec = []
    for s in secs:
        orig = orig_by_docref.get(s["corr"] or "")
        if orig is not None and orig.get("rec_jur_code") and s["rec"] != orig["rec_jur_code"]:
            bad_rec.append(f"{s['key']} ({s['rec']} vs {orig['rec_jur_code']})")
    check("RecJurCode equals the original section's RecJurCode (rule 98200)",
          not bad_rec, "; ".join(bad_rec))

    return results


def build_merged_view(original_ctx: dict, correction_xml_str: str) -> str:
    """Splice the corrected sections into the original report — the state ESTV's
    database will hold after acceptance. ESTV evaluates the cross-element rules
    (98201, 70008 ff.) against THIS, so we re-run the full structural validation
    on it (Layer 2). Critical for filings with no test-portal access."""
    merged = ET.fromstring(ET.tostring(original_ctx["root"]))
    mbody  = merged.find(N + "GLOBEBody")
    pos_by_docref = {}
    for el in mbody:
        ds = el.find(N + "DocSpec")
        if ds is not None and ds.find(S + "DocRefId") is not None:
            pos_by_docref[(ds.find(S + "DocRefId").text or "").strip()] = el

    corr_body = ET.fromstring(correction_xml_str).find(N + "GLOBEBody")
    for _key, el in _walk_sections(corr_body):
        ds   = el.find(N + "DocSpec")
        dti  = (ds.find(S + "DocTypeIndic").text or "").strip() if ds.find(S + "DocTypeIndic") is not None else ""
        corr = ds.find(S + "CorrDocRefId")
        target = pos_by_docref.get((corr.text or "").strip()) if corr is not None and corr.text else None
        if target is None:
            continue                     # dangling CorrDocRefId → Layer 1 already fails it
        if dti in ("OECD3", "OECD13"):   # Storno → section disappears from the report
            mbody.remove(target)
        else:                            # correction → whole-element replacement
            idx = list(mbody).index(target)
            mbody.remove(target)
            mbody.insert(idx, ET.fromstring(ET.tostring(el)))
    ET.indent(merged, space="  ")
    return f"<?xml version='1.0' encoding='utf-8'?>\n{ET.tostring(merged, encoding='unicode')}"


def validate_correction(correction_xml: str, original_ctx: dict):
    """Both layers: (correction-file checks, merged-state structural checks)."""
    layer1 = validate_correction_xml(correction_xml, original_ctx)
    try:
        layer2 = validate_xml(build_merged_view(original_ctx, correction_xml))
    except Exception as e:  # merged view must never crash the app
        layer2 = [("Merged view could not be built", False, str(e))]
    return layer1, layer2


# ─── TRANSLATIONS ────────────────────────────────────────────────────────────

T: dict[str, dict[str, str]] = {
    "step1":               {"EN": "1. Upload GIR template",            "DE": "1. GIR-Vorlage hochladen"},
    "upload_label":        {"EN": "GIR Template (.xlsx or .xlsm)", "DE": "GIR-Vorlage (.xlsx oder .xlsm)"},
    "upload_help":         {"EN": 'OECD GIR template with sheets "1 MNE Group Information" and "3 GloBE Computations"',
                            "DE": 'OECD-GIR-Vorlage mit Tabellenblättern "1 MNE Group Information" und "3 GloBE Computations"'},
    "autofilled":          {"EN": "Company details auto-filled from the template — review and adjust if needed.",
                            "DE": "Unternehmensangaben aus der Vorlage übernommen — bitte prüfen und ggf. anpassen."},
    "autofill_failed":     {"EN": "Could not auto-fill from the template ({}). Enter details manually below.",
                            "DE": "Automatisches Ausfüllen aus der Vorlage fehlgeschlagen ({}). Bitte Angaben unten manuell erfassen."},
    "ce_count":            {"EN": "{} constituent entities read from the corporate structure (sheet 1).",
                            "DE": "{} Untereinheiten aus der Konzernstruktur (Blatt 1) gelesen."},
    "ce_table_title":      {"EN": "Constituent entities — TINs", "DE": "Untereinheiten — UID/TIN"},
    "ce_table_help":       {"EN": "Fill any missing TIN here (ESTV requires a TIN for every entity). Edits override the Excel.",
                            "DE": "Fehlende UID/TIN hier ergänzen (die ESTV verlangt eine TIN je Einheit). Änderungen überschreiben die Excel-Datei."},
    "ce_col_entity":       {"EN": "Entity",        "DE": "Einheit"},
    "ce_col_jur":          {"EN": "Jurisdiction",  "DE": "Jurisdiktion"},
    "ce_col_tin":          {"EN": "TIN",           "DE": "UID/TIN"},
    "ce_missing_warn":     {"EN": "⚠️ {} entity(ies) still without a TIN — ESTV will reject these (error 70006). Fill them above before submitting.",
                            "DE": "⚠️ {} Einheit(en) noch ohne UID/TIN — die ESTV weist diese zurück (Fehler 70006). Bitte oben ergänzen."},
    "ce_all_have_tin":     {"EN": "✅ All constituent entities have a TIN.", "DE": "✅ Alle Untereinheiten haben eine UID/TIN."},
    "sh_title":            {"EN": "Safe-harbour jurisdictions", "DE": "Safe-Harbour-Jurisdiktionen"},
    "sh_help":             {"EN": "{} jurisdiction(s) elect a safe harbour (from the 'Safe Harbours' tabs). Each becomes a Summary + JurisdictionSection. Empty tabs are ignored.",
                            "DE": "{} Jurisdiktion(en) wählen einen Safe Harbour (aus den 'Safe Harbours'-Blättern). Jede wird zu Summary + JurisdictionSection. Leere Blätter werden ignoriert."},
    "step2":               {"EN": "2. Review company details",         "DE": "2. Unternehmensangaben prüfen"},
    "company_name":        {"EN": "Company name",                      "DE": "Firmenname"},
    "tin":                 {"EN": "TIN",                               "DE": "UID/TIN"},
    "tin_issued_by":       {"EN": "TIN issued by (ISO 3166-1 Alpha-2)","DE": "TIN ausgestellt von (ISO 3166-1 Alpha-2)"},
    "jurisdiction":        {"EN": "Jurisdiction (ISO 3166-1 Alpha-2)", "DE": "Jurisdiktion (ISO 3166-1 Alpha-2)"},
    "currency":            {"EN": "Currency (ISO 4217)",               "DE": "Währung (ISO 4217)"},
    "fas_label":           {"EN": "Financial Accounting Standard",     "DE": "Rechnungslegungsstandard"},
    "period_start":        {"EN": "Period start",                      "DE": "Periodenstart"},
    "period_end":          {"EN": "Period end",                        "DE": "Periodenende"},
    "partner_country":     {"EN": "Partner country (RecJurCode)",      "DE": "Partnerstaat (RecJurCode)"},
    "partner_help":        {"EN": "ISO 3166-1 Alpha-2 country code of the partner jurisdiction (must not be CH)",
                            "DE": "ISO 3166-1 Alpha-2 Ländercode des Partnerstaats (darf nicht CH sein)"},
    "advanced":            {"EN": "Advanced options",                  "DE": "Erweiterte Optionen"},
    "filing_role":         {"EN": "Filing role",                       "DE": "Einreichungsrolle"},
    "filing_role_help":    {"EN": "Role as registered in the ESTV ePortal",
                            "DE": "Rolle gemäss Registrierung im ESTV ePortal"},
    "tin_type_label":      {"EN": "TIN type",                         "DE": "TIN-Typ"},
    "tin_type_help":       {"EN": "Type of identifier used as TIN",   "DE": "Art des als TIN verwendeten Identifikators"},
    "cfs_upe":             {"EN": "CFS of UPE",                       "DE": "CFS der UPE"},
    "cfs_upe_help":        {"EN": "Type of Consolidated Financial Statement of the Ultimate Parent Entity",
                            "DE": "Art des konsolidierten Abschlusses der obersten Muttergesellschaft"},
    "submission_mode":     {"EN": "Submission mode",                   "DE": "Einreichungsmodus"},
    "mode_production":     {"EN": "Production (OECD1)",                "DE": "Produktion (OECD1)"},
    "mode_test":           {"EN": "Test / CTS (OECD11)",                     "DE": "Test / CTS (OECD11)"},
    "mode_help":           {"EN": "Use Test/CTS for the acceptance portal (eportal-a.admin.ch). Use Production for the live portal (eportal.admin.ch).",
                            "DE": "Test/CTS für das Abnahmeportal (eportal-a.admin.ch), Produktion für das Live-Portal (eportal.admin.ch)."},
    "mode_test_banner":    {"EN": "⚠️ TEST mode (OECD11) — this file is for the acceptance portal (eportal-a.admin.ch) only. Do NOT upload it to the live production portal. Switch to Production (OECD1) before the real filing.",
                            "DE": "⚠️ TEST-Modus (OECD11) — diese Datei ist nur für das Abnahmeportal (eportal-a.admin.ch). NICHT ins Live-Produktionsportal hochladen. Vor der echten Einreichung auf Produktion (OECD1) wechseln."},
    "testmode_badge":      {"EN": "Testmode",                            "DE": "Testmodus"},
    "gir401":              {"EN": "GIR401 — Ultimate Parent Entity (UPE)",    "DE": "GIR401 — Oberste Muttergesellschaft (UPE)"},
    "gir402":              {"EN": "GIR402 — Designated Filing Entity (DFE)",  "DE": "GIR402 — Benannte Einreichungsstelle (DFE)"},
    "gir404":              {"EN": "GIR404 — Constituent Entity (CE)",         "DE": "GIR404 — Untereinheit (CE)"},
    "gir3001":             {"EN": "GIR3001 — Tax Identification Number (TIN)","DE": "GIR3001 — Steueridentifikationsnummer (TIN)"},
    "gir3002":             {"EN": "GIR3002 — Functional equivalent",          "DE": "GIR3002 — Funktionales Äquivalent"},
    "gir501":              {"EN": "GIR501 — Consolidated Financial Statement (subparagraph a)",
                            "DE": "GIR501 — Konsolidierter Abschluss (Buchstabe a)"},
    "gir502":              {"EN": "GIR502 — Combined Financial Statement (subparagraph b)",
                            "DE": "GIR502 — Kombinierter Abschluss (Buchstabe b)"},
    "gir503":              {"EN": "GIR503 — Other",                           "DE": "GIR503 — Sonstiges"},
    "step3":               {"EN": "3. Validate file",                  "DE": "3. Datei validieren"},
    "generate_btn":        {"EN": "Validate file",                     "DE": "Datei validieren"},
    "upload_first":        {"EN": "Please upload an Excel file first.","DE": "Bitte laden Sie zuerst eine Excel-Datei hoch."},
    "spinner_gen":         {"EN": "Reading Excel and building XML…",   "DE": "Excel wird gelesen und XML wird erstellt…"},
    "upload_to_enable":    {"EN": "Upload the Excel file above to enable export.",
                            "DE": "Laden Sie die Excel-Datei oben hoch, um den Export zu aktivieren."},
    "validation_title":    {"EN": "Structural validation",             "DE": "Strukturelle Validierung"},
    "checks_passed":       {"EN": "checks passed",                     "DE": "Prüfungen bestanden"},
    "fix_issues":          {"EN": ("Fix the issues above, then re-generate. "
                                   "Once all checks pass, validate against the official ESTV XSD before submission."),
                            "DE": ("Beheben Sie die oben genannten Probleme und generieren Sie erneut. "
                                   "Sobald alle Prüfungen bestanden sind, validieren Sie gegen das offizielle ESTV XSD vor der Einreichung.")},
    "all_passed":          {"EN": "All structural checks passed.",     "DE": "Alle strukturellen Prüfungen bestanden."},
    "download_blocked":    {"EN": "Not all checks pass — fix the issues above and re-validate. The downloads stay disabled in Step 4.",
                            "DE": "Nicht alle Prüfungen bestanden — beheben Sie die Probleme oben und validieren Sie erneut. Die Downloads bleiben in Schritt 4 deaktiviert."},
    "encrypt_blocked":     {"EN": "The downloads are disabled until all structural checks pass (Step 3).",
                            "DE": "Die Downloads sind deaktiviert, bis alle strukturellen Prüfungen bestanden sind (Schritt 3)."},
    "download_xml":        {"EN": "Download raw XML",                  "DE": "Roh-XML herunterladen"},
    "preview_xml":         {"EN": "Preview XML",                       "DE": "XML-Vorschau"},
    "sheet_not_found":     {"EN": 'Required sheet(s) not found: {}. Use the OECD GIR template with "1 MNE Group Information" and "3 GloBE Computations".',
                            "DE": 'Erforderliche(s) Tabellenblatt/-blätter nicht gefunden: {}. Verwenden Sie die OECD-GIR-Vorlage mit "1 MNE Group Information" und "3 GloBE Computations".'},
    "error_msg":           {"EN": "Error: {}",                         "DE": "Fehler: {}"},
    "step4":               {"EN": "4. Download for ESTV",             "DE": "4. Download für die ESTV"},
    "step4_caption":       {"EN": ("Upload the ESTV public key (ESTV-PublicKey.pem) from the myESTV portal. "
                                   "The app will produce an encrypted .zip ready to upload directly to the GIR-Applikation."),
                            "DE": ("Laden Sie den öffentlichen ESTV-Schlüssel (ESTV-PublicKey.pem) aus dem myESTV-Portal hoch. "
                                   "Die App erstellt eine verschlüsselte .zip-Datei, die direkt in die GIR-Applikation hochgeladen werden kann.")},
    "pem_label":           {"EN": "ESTV Public Key (.pem)",            "DE": "Öffentlicher ESTV-Schlüssel (.pem)"},
    "encrypt_btn":         {"EN": "Encrypt & download",               "DE": "Verschlüsseln & herunterladen"},
    "generate_first":      {"EN": "Generate the XML first (Step 3).", "DE": "Generieren Sie zuerst das XML (Schritt 3)."},
    "upload_pem":          {"EN": "Upload the ESTV public key (.pem) above.",
                            "DE": "Laden Sie oben den öffentlichen ESTV-Schlüssel (.pem) hoch."},
    "spinner_enc":         {"EN": "Encrypting…",                      "DE": "Verschlüsselung läuft…"},
    "download_zip":        {"EN": "⬇️  Download encrypted ZIP",       "DE": "⬇️  Verschlüsselte ZIP herunterladen"},
    "ready_estv":          {"EN": "Ready to upload to myESTV → GIR-Applikation.",
                            "DE": "Bereit zum Hochladen in myESTV → GIR-Applikation."},
    "enc_failed":          {"EN": "Encryption failed: {}",            "DE": "Verschlüsselung fehlgeschlagen: {}"},
    "gen_first_long":      {"EN": "Generate the XML in Step 3 first, then encrypt here.",
                            "DE": "Generieren Sie zuerst das XML in Schritt 3, dann verschlüsseln Sie hier."},
    "bundled_key_info":    {"EN": "Using the bundled ESTV public key — the download is an encrypted .zip ready to upload to myESTV → GIR-Applikation.",
                            "DE": "Mit dem integrierten öffentlichen ESTV-Schlüssel — der Download ist eine verschlüsselte .zip, bereit zum Hochladen in myESTV → GIR-Applikation."},
    "override_key":        {"EN": "Use a different key",              "DE": "Anderen Schlüssel verwenden"},
    "override_active":     {"EN": "Custom key active",                "DE": "Eigener Schlüssel aktiv"},
    "err_jurisdiction":    {"EN": "Jurisdiction must be exactly 2 uppercase letters (e.g. CH)",
                            "DE": "Die Jurisdiktion muss genau 2 Grossbuchstaben sein (z.B. CH)"},
    "err_currency":        {"EN": "Currency must be exactly 3 uppercase letters (e.g. CHF)",
                            "DE": "Die Währung muss genau 3 Grossbuchstaben sein (z.B. CHF)"},
    "err_period_start":    {"EN": "Period start must be YYYY-MM-DD",   "DE": "Periodenstart muss im Format JJJJ-MM-TT sein"},
    "err_period_end":      {"EN": "Period end must be YYYY-MM-DD",     "DE": "Periodenende muss im Format JJJJ-MM-TT sein"},
    "err_company":         {"EN": "Company name is required",          "DE": "Firmenname ist erforderlich"},
    "err_tin":             {"EN": "TIN is required",                   "DE": "UID/TIN ist erforderlich"},
    "err_tin_issued":      {"EN": "TIN issued by must be exactly 2 uppercase letters (e.g. CH)",
                            "DE": "TIN ausgestellt von muss genau 2 Grossbuchstaben sein (z.B. CH)"},
    "err_rec_jur":         {"EN": "Partner country (RecJurCode) must be exactly 2 uppercase letters (e.g. DE)",
                            "DE": "Partnerstaat (RecJurCode) muss genau 2 Grossbuchstaben sein (z.B. DE)"},
    "upe_rules_label":     {"EN": "UPE GloBE Rules",                  "DE": "UPE GloBE-Regeln"},
    "upe_rules_help":      {"EN": "GloBE rules applicable to the Ultimate Parent Entity (GIR201–GIR205). Use GIR204 for QDMTT filers.",
                            "DE": "Anwendbare GloBE-Regeln für die oberste Muttergesellschaft (GIR201–GIR205). GIR204 für QDMTT-Einreicher."},
    "upe_globe_status_label": {"EN": "UPE GloBE Status",            "DE": "UPE GloBE-Status"},
    "upe_globe_status_help":  {"EN": "GloBE entity classification of the Ultimate Parent Entity (GIR301–GIR318). GIR301 for standard Constituent Entity.",
                               "DE": "GloBE-Entitätsstatus der obersten Muttergesellschaft (GIR301–GIR318). GIR301 für normale Untereinheit."},
    "in_coop":             {"EN": "in cooperation with",             "DE": "in Zusammenarbeit mit"},
    "summary_label":       {"EN": "Plain Language Summary",            "DE": "Verständliche Zusammenfassung"},
    "sum_filing":          {"EN": "Filing Information",               "DE": "Einreichungsangaben"},
    "sum_company":         {"EN": "Company",                          "DE": "Unternehmen"},
    "sum_tin":             {"EN": "TIN",                              "DE": "UID/TIN"},
    "sum_jurisdiction":    {"EN": "Jurisdiction",                     "DE": "Jurisdiktion"},
    "sum_period":          {"EN": "Reporting Period",                 "DE": "Berichtsperiode"},
    "sum_currency":        {"EN": "Currency",                         "DE": "Währung"},
    "sum_fas":             {"EN": "Accounting Standard",              "DE": "Rechnungslegungsstandard"},
    "sum_role":            {"EN": "Filing Role",                      "DE": "Einreichungsrolle"},
    "sum_partner":         {"EN": "Partner Country",                  "DE": "Partnerstaat"},
    "sum_income":          {"EN": "GloBE Income Computation",        "DE": "GloBE-Einkommensberechnung"},
    "sum_fanil":           {"EN": "Adjusted FANIL",                   "DE": "Angepasstes FANIL"},
    "sum_income_adj":      {"EN": "Income Adjustments (non-zero)",   "DE": "Einkommensanpassungen (ungleich null)"},
    "sum_net_income":      {"EN": "Net GloBE Income",                "DE": "Netto GloBE-Einkommen"},
    "sum_tax":             {"EN": "Covered Tax Computation",         "DE": "Berechnung der anrechenbaren Steuer"},
    "sum_curr_tax":        {"EN": "Aggregate Current Tax",           "DE": "Aggregierte laufende Steuer"},
    "sum_tax_adj":         {"EN": "Tax Adjustments (non-zero)",      "DE": "Steueranpassungen (ungleich null)"},
    "sum_adj_tax":         {"EN": "Adjusted Covered Tax",            "DE": "Angepasste anrechenbare Steuer"},
    "sum_result":          {"EN": "Result",                           "DE": "Ergebnis"},
    "sum_etr":             {"EN": "Effective Tax Rate (ETR)",        "DE": "Effektiver Steuersatz (ETR)"},
    "disclaimer_label":    {"EN": "Disclaimer",                        "DE": "Haftungsausschluss"},
    "disclaimer_text":     {
        "EN": (
            "This tool converts the figures and entity data from your GIR Excel template into the OECD GloBE "
            "Information Return XML format and encrypts it for the ESTV portal. It does not calculate, interpret "
            "or verify any tax position — the output reflects only the data you enter, and the completeness and "
            "accuracy of the return remain your responsibility. The tool runs structural and schema checks, but "
            "these do not replace ESTV's own validation: always submit to the ESTV test (ABN) portal first and "
            "review the returned status message before filing for real. This tool does not constitute legal or "
            "tax advice and is provided without warranty; have the return reviewed by a qualified professional."
        ),
        "DE": (
            "Dieses Tool wandelt die Zahlen und Einheitsdaten aus Ihrer GIR-Excel-Vorlage in das OECD-GloBE-"
            "Information-Return-XML-Format um und verschlüsselt es für das ESTV-Portal. Es berechnet, interpretiert "
            "oder prüft keine steuerliche Beurteilung — der Output gibt ausschliesslich die von Ihnen erfassten "
            "Daten wieder, und die Vollständigkeit und Richtigkeit der Meldung liegen in Ihrer Verantwortung. Das "
            "Tool führt strukturelle und Schema-Prüfungen durch, diese ersetzen jedoch nicht die Validierung der "
            "ESTV: Reichen Sie stets zuerst im ESTV-Testportal (ABN) ein und prüfen Sie die zurückgegebene "
            "Statusmeldung, bevor Sie produktiv einreichen. Dieses Tool stellt keine Rechts- oder Steuerberatung "
            "dar und wird ohne Gewähr bereitgestellt; lassen Sie die Meldung von einer Fachperson prüfen."
        ),
    },
    # ── Correction mode (Rektifikat) ──
    "filing_type":         {"EN": "Filing type",                       "DE": "Meldungsart"},
    "mode_neumeldung":     {"EN": "New filing (Neumeldung)",           "DE": "Neumeldung"},
    "mode_korrektur":      {"EN": "Correction (Rektifikat)",           "DE": "Korrektur (Rektifikat)"},
    "korr_badge":          {"EN": "Correction",                        "DE": "Korrektur"},
    "korr_upload_label":   {"EN": "Originally submitted XML (raw)",    "DE": "Original eingereichtes XML (Roh-XML)"},
    "korr_upload_help":    {"EN": ("The RAW XML of the last accepted filing (from 'Download raw XML'). The _encrypted.zip "
                                   "cannot be used — it is encrypted with ESTV's public key. If a correction was already "
                                   "accepted before, upload that LATEST correction instead (correction chains)."),
                            "DE": ("Das ROH-XML der letzten akzeptierten Meldung ('Roh-XML herunterladen'). Die _encrypted.zip "
                                   "kann nicht verwendet werden — sie ist mit dem öffentlichen ESTV-Schlüssel verschlüsselt. "
                                   "Wurde bereits eine Korrektur akzeptiert, laden Sie stattdessen DIESE letzte Korrektur hoch (Korrekturketten).")},
    "korr_need_both":      {"EN": "Correction mode: upload the corrected Excel template AND the originally submitted raw XML (Step 1).",
                            "DE": "Korrekturmodus: Laden Sie die korrigierte Excel-Vorlage UND das original eingereichte Roh-XML hoch (Schritt 1)."},
    "korr_parse_ok":       {"EN": "Original parsed: {} sections, MessageRefId {}.",
                            "DE": "Original eingelesen: {} Berichtselemente, MessageRefId {}."},
    "korr_parse_err":      {"EN": "Could not read the original XML: {}", "DE": "Original-XML konnte nicht gelesen werden: {}"},
    "korr_sections_title": {"EN": "Sections to correct",               "DE": "Zu korrigierende Berichtselemente"},
    "korr_sections_help":  {"EN": ("Auto-comparison of your corrected Excel against the original filing. Only ticked sections go into "
                                   "the correction message — unchanged sections are omitted (ESTV rule). Review with your tax expert."),
                            "DE": ("Automatischer Vergleich Ihrer korrigierten Excel-Datei mit der Originalmeldung. Nur angehakte "
                                   "Berichtselemente kommen in die Korrekturmeldung — unveränderte werden weggelassen (ESTV-Regel). "
                                   "Mit Ihrer Steuerfachperson prüfen.")},
    "korr_st_changed":     {"EN": "CHANGED",                           "DE": "GEÄNDERT"},
    "korr_st_unchanged":   {"EN": "unchanged",                         "DE": "unverändert"},
    "korr_st_new":         {"EN": "NEW — not allowed here",            "DE": "NEU — hier nicht erlaubt"},
    "korr_st_missing":     {"EN": "missing in Excel",                  "DE": "fehlt im Excel"},
    "korr_new_blocked":    {"EN": ("{} new section(s) found that are not in the original filing. New sections cannot be part of a "
                                   "correction (GIR102) — they need a separate Neumeldung (GIR101) with the FilingInfo resent. "
                                   "They will be EXCLUDED from this correction file."),
                            "DE": ("{} neue(s) Berichtselement(e) gefunden, die nicht in der Originalmeldung sind. Neue Elemente dürfen "
                                   "nicht in eine Korrektur (GIR102) — sie brauchen eine separate Neumeldung (GIR101) mit erneut "
                                   "gesendeter FilingInfo. Sie werden aus dieser Korrekturdatei AUSGESCHLOSSEN.")},
    "korr_advanced":       {"EN": "Advanced — DocRefIds & Storno",     "DE": "Erweitert — DocRefIds & Storno"},
    "korr_adv_help":       {"EN": ("DocRefIds are pre-filled from the uploaded original. Only override them if you must recover them "
                                   "from an ESTV status message instead. Storno (deletion) removes a section from the report — "
                                   "it CANNOT be undone by a later correction."),
                            "DE": ("DocRefIds sind aus dem hochgeladenen Original vorbefüllt. Nur überschreiben, wenn sie stattdessen "
                                   "aus einer ESTV-Statusmeldung rekonstruiert werden müssen. Storno löscht ein Berichtselement aus "
                                   "dem Bericht — das kann durch eine spätere Korrektur NICHT rückgängig gemacht werden.")},
    "korr_fi_docref":      {"EN": "FilingInfo DocRefId (original)",    "DE": "FilingInfo DocRefId (Original)"},
    "korr_storno_label":   {"EN": "Storno (delete) {}",                "DE": "Storno (löschen) {}"},
    "korr_none_selected":  {"EN": "Select at least one section to correct or delete.",
                            "DE": "Wählen Sie mindestens ein Berichtselement zum Korrigieren oder Löschen aus."},
    "korr_l1_title":       {"EN": "Correction-file checks (ESTV rules)", "DE": "Prüfungen der Korrekturdatei (ESTV-Regeln)"},
    "korr_l2_title":       {"EN": "Merged report after correction — full re-validation",
                            "DE": "Gesamtbericht nach Korrektur — vollständige Neuvalidierung"},
    "korr_summary":        {"EN": "Correction message: {} section(s) corrected, {} deleted — unchanged sections omitted.",
                            "DE": "Korrekturmeldung: {} Berichtselement(e) korrigiert, {} storniert — unveränderte weggelassen."},
}


# ─── STREAMLIT UI ────────────────────────────────────────────────────────────

MME_LOGO_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mme_logo.svg")
ESTV_PEM_PATH   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "estv-publickey.pem")
MUTARA_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mutara_logo.png")

def _load_mutara_logo_b64() -> str | None:
    try:
        with open(MUTARA_LOGO_PATH, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except OSError:
        return None

_MUTARA_LOGO_B64 = _load_mutara_logo_b64()

def _load_estv_pem() -> bytes | None:
    try:
        with open(ESTV_PEM_PATH, "rb") as f:
            return f.read()
    except OSError:
        return None

_BUNDLED_PEM = _load_estv_pem()

def _load_mme_logo_svg():
    try:
        with open(MME_LOGO_PATH, "r") as f:
            return f.read()
    except OSError:
        return None

_MME_LOGO_SVG = _load_mme_logo_svg()

_favicon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "favicon.png")
_favicon = Image.open(_favicon_path) if os.path.exists(_favicon_path) else "🌐"

st.set_page_config(
    page_title="GloBE XML Export | MME",
    page_icon=_favicon,
    layout="centered",
)

# ── Language state ────────────────────────────────────────────────────────────
if "lang" not in st.session_state:
    st.session_state["lang"] = "EN"
lang = st.session_state["lang"]

st.markdown("""
<style>
    /* ── MME GIR — design tokens ported from the GIR Portal (portal.css) ───── */
    :root {
        --red: #e0303c; --red-dark: #c0272f; --red-tint: #fdeef0;
        --ink: #1f262e; --slate: #5a6672; --slate-light: #8a96a3;
        --line: #e4e9ef; --line-strong: #d2dae3;
        --bg: #f6f8fb; --surface: #ffffff;
        --green: #1f9d57; --green-tint: #e9f7ef;
        --shadow-sm: 0 1px 2px rgba(31,38,46,.06), 0 1px 3px rgba(31,38,46,.04);
        --shadow-md: 0 4px 16px rgba(31,38,46,.08), 0 2px 6px rgba(31,38,46,.05);
        --radius: 12px; --radius-sm: 8px;
        --font: "Helvetica Neue", Helvetica, "Segoe UI", Arial, sans-serif;
    }
    html, body, [class*="css"], .stApp { font-family: var(--font); color: var(--ink); }

    /* App canvas: calm light surface with the portal's subtle radial wash */
    .stApp, [data-testid="stAppViewContainer"] {
        background: var(--bg);
        background-image:
            radial-gradient(900px 500px at 88% -8%, rgba(224,48,60,.06), transparent 60%),
            radial-gradient(700px 500px at -5% 0%, rgba(31,38,46,.035), transparent 55%);
        background-attachment: fixed;
    }
    .block-container { max-width: 880px; padding-top: 5.4rem; padding-bottom: 5rem; }

    /* Hide Streamlit chrome for a product feel */
    #MainMenu, footer, [data-testid="stToolbar"], [data-testid="stDecoration"],
    [data-testid="stStatusWidget"] { display: none !important; }
    [data-testid="stHeader"] { background: transparent; height: 0; pointer-events: none; }

    /* ── step cards: st.container(border=True, key="girc*") → portal .step ──── */
    /* Streamlit 1.57 renders the bordered container as a transparent stVerticalBlock;
       the stable st-key-* class lets us paint it white like the portal card. */
    .stApp [class*="st-key-girc"] {
        background: var(--surface) !important;
        border: 1px solid var(--line) !important;
        border-radius: var(--radius) !important;
        box-shadow: var(--shadow-sm) !important;
        padding: 22px 26px !important;
        margin-bottom: 18px;
    }

    /* ── numbered step header (.step-head / .step-num / .step-title) ───────── */
    .step-head { display: flex; align-items: center; gap: 13px; margin: 2px 0 16px; }
    .step-num {
        width: 30px; height: 30px; flex: none; border-radius: 50%;
        background: var(--red); color: #fff; font-weight: 700; font-size: .9rem;
        display: grid; place-items: center; box-shadow: 0 2px 6px rgba(224,48,60,.3);
    }
    .step-title { margin: 0; font-size: 1.06rem; font-weight: 700;
        letter-spacing: -.01em; color: var(--ink); }

    /* ── language toggle (radio → pills), aligned right ───────────────────── */
    div[data-testid="stRadio"] label { display: none; }
    div[data-testid="stRadio"] > div { display: flex; gap: 4px; justify-content: flex-end; }
    div[data-testid="stRadio"] > div > label {
        display: flex !important; padding: 4px 12px;
        border: 1px solid var(--line-strong); border-radius: 20px;
        font-size: .76rem; font-weight: 600; color: var(--slate); cursor: pointer;
        transition: all .15s ease;
    }
    div[data-testid="stRadio"] > div > label:hover { border-color: var(--red); color: var(--red); }
    div[data-testid="stRadio"] > div > label[data-checked="true"] {
        background: var(--red); border-color: var(--red); color: #fff;
    }

    /* ── buttons: portal pill buttons with shadow + disabled state ────────── */
    .stButton > button[kind="primary"], .stDownloadButton > button {
        background: var(--red) !important; border: none !important; color: #fff !important;
        font-weight: 600 !important; border-radius: 30px !important;
        padding: .55rem 1.5rem !important; box-shadow: 0 2px 8px rgba(224,48,60,.28) !important;
        transition: background .15s ease, transform .08s ease !important;
    }
    .stButton > button[kind="primary"]:hover, .stDownloadButton > button:hover {
        background: var(--red-dark) !important;
    }
    .stButton > button[kind="primary"]:active, .stDownloadButton > button:active {
        transform: translateY(1px);
    }
    .stButton > button[kind="primary"]:disabled, .stDownloadButton > button:disabled {
        background: #d7dde4 !important; color: #fff !important; box-shadow: none !important;
        cursor: not-allowed; opacity: 1 !important;
    }
    .stDownloadButton > button { width: 100%; justify-content: center; }

    /* ── inputs / selects: white surface + portal focus ring ──────────────── */
    .stTextInput [data-baseweb="base-input"], .stTextInput input,
    .stDateInput [data-baseweb="base-input"], .stDateInput input,
    .stNumberInput [data-baseweb="base-input"],
    .stSelectbox [data-baseweb="select"] > div,
    .stMultiSelect [data-baseweb="select"] > div {
        background: var(--surface) !important;
        border-radius: var(--radius-sm) !important;
    }
    .stTextInput [data-baseweb="base-input"], .stDateInput [data-baseweb="base-input"],
    .stSelectbox [data-baseweb="select"] > div {
        border: 1px solid var(--line-strong) !important;
    }
    .stTextInput input:focus, .stDateInput input:focus {
        border-color: var(--red) !important; box-shadow: 0 0 0 3px rgba(224,48,60,.12) !important;
    }

    /* ── file uploader → portal .dropzone (centered, icon, custom copy) ────── */
    [data-testid="stFileUploaderDropzone"] {
        display: flex !important; flex-direction: column !important;
        align-items: center !important; justify-content: center !important;
        gap: 8px; text-align: center; min-height: 150px;
        border: 2px dashed var(--line-strong) !important; border-radius: var(--radius) !important;
        background: #fbfcfe !important; padding: 28px 20px !important; transition: all .18s ease;
    }
    [data-testid="stFileUploaderDropzone"]:hover {
        border-color: var(--red) !important; background: var(--red-tint) !important;
    }
    [data-testid="stFileUploaderDropzone"]::before {
        content: "📄"; font-size: 2rem; line-height: 1; order: 0;
    }
    /* replace Streamlit's size line with the portal headline */
    [data-testid="stFileUploaderDropzoneInstructions"] { order: 1; }
    [data-testid="stFileUploaderDropzoneInstructions"] span { font-size: 0 !important; }
    [data-testid="stFileUploaderDropzoneInstructions"] span::after {
        content: "Drop your GIR template here, or click to browse";
        font-size: .95rem; font-weight: 600; color: var(--ink);
    }
    [data-testid="stFileUploaderDropzone"] button {
        order: 2; margin-top: 2px;
        background: var(--surface) !important; color: var(--red-dark) !important;
        border: 1px solid var(--line-strong) !important; border-radius: 20px !important;
    }
    [data-testid="stFileUploaderDropzone"] button:hover {
        border-color: var(--red) !important; background: var(--red-tint) !important;
    }

    /* ── metrics / validation / misc ──────────────────────────────────────── */
    [data-testid="stMetricValue"] { color: var(--red) !important; font-weight: 700; font-size: 1.05rem !important; }
    [data-testid="stMetricLabel"] { font-size: .72rem !important; text-transform: uppercase;
        letter-spacing: .05em; color: var(--slate-light) !important; }
    [data-testid="stExpander"] details { border: 1px solid var(--line) !important;
        border-radius: var(--radius-sm) !important; }
    hr { border-color: var(--line) !important; }
    [data-testid="stCaptionContainer"], .stCaption { color: var(--slate-light) !important; }
    a { color: var(--red-dark); }

    /* ── header bar: portal .site-header — FULL-WIDTH fixed white bar ───────── */
    /* Pure-HTML bar (no Streamlit columns inside → layout can't break); the real
       EN/DE radio is pinned into its right side via the st-key-girlang container. */
    .gir-topbar {
        position: fixed; top: 0; left: 0; right: 0; z-index: 100;
        background: var(--surface); border-bottom: 1px solid var(--line);
        box-shadow: var(--shadow-sm); padding: 9px 0;
        pointer-events: none;   /* let clicks pass through to the pinned EN/DE toggle */
    }
    .gir-topbar-inner {
        max-width: 880px; margin: 0 auto; padding: 0 8px;
        display: flex; align-items: center; gap: 22px;
    }
    .stApp [class*="st-key-girlang"] {
        position: fixed; top: 11px; z-index: 2147483000;
        right: max(12px, calc((100vw - 880px) / 2 + 8px));
        width: auto !important;
    }
    .gir-title {
        margin: 0; color: var(--ink); font-size: 1.15rem; font-weight: 700;
        letter-spacing: -.01em; font-family: var(--font); white-space: nowrap;
        border-left: 1px solid var(--line-strong); padding-left: 22px;
    }
</style>
""", unsafe_allow_html=True)


def step_header(num: int, title: str, badge: str | None = None) -> None:
    """Render a portal-style numbered step header (red circle + title).
    Strips a leading 'N. ' from the translated section title to avoid a double number.
    `badge` (e.g. 'Testmode') renders an amber ⚠️ pill next to the title."""
    clean = re.sub(r"^\s*\d+\.\s*", "", title)
    badge_html = (
        f"<span style='margin-left:10px; font-size:0.72rem; font-weight:600; "
        f"color:#9a6700; background:#fff8e1; border:1px solid #f3d27a; "
        f"border-radius:6px; padding:1px 8px; vertical-align:middle; "
        f"white-space:nowrap;'>⚠️ {badge}</span>"
        if badge else ""
    )
    st.markdown(
        f"<div class='step-head'><div class='step-num'>{num}</div>"
        f"<div class='step-title'>{clean}{badge_html}</div></div>",
        unsafe_allow_html=True,
    )


def _card_open(key: str):
    """Open a bordered container (portal .step card) and make it the active target
    without indenting the section body. The key yields a stable st-key-<key> class
    that the CSS paints white. Closed with _card_close()."""
    c = st.container(border=True, key=key)
    c.__enter__()
    return c


def _card_close(c) -> None:
    c.__exit__(None, None, None)

# ── Header row: logo + title | language toggle ────────────────────────────────
if _MME_LOGO_SVG:
    scaled_svg = _MME_LOGO_SVG.replace(
        'width="204" height="65"',
        'width="90" height="29" viewBox="0 0 204 65"',
    )
else:
    scaled_svg = ""

# Full-width fixed top bar — pure HTML (no Streamlit columns inside, so the
# language toggle can't disturb it). The real EN/DE radio below is CSS-pinned in.
st.markdown(
    f"""<div class='gir-topbar'><div class='gir-topbar-inner'>
        <div style='display:flex; flex-direction:column; align-items:flex-start;'>
            <div>{scaled_svg}</div>
            <span style='font-size:0.64rem; color:#8a96a3; letter-spacing:0.04em; margin-top:2px;'>v{VERSION}</span>
        </div>
        <div class='gir-title'>GloBE Information Return (GIR)</div>
    </div></div>""",
    unsafe_allow_html=True,
)

_lang = st.container(key="girlang")
with _lang:
    selected = st.radio(
        "lang", ["EN", "DE"],
        index=0 if lang == "EN" else 1,
        horizontal=True,
        label_visibility="collapsed",
        key="lang_radio",
    )
    if selected != lang:
        st.session_state["lang"] = selected
        st.rerun()

# ── Step 1: Upload ────────────────────────────────────────────────────────────
_card1 = _card_open("girc1")
step_header(1, T["step1"][lang])
# The global stRadio CSS hides radio labels (language-pill styling) → render the
# label as markdown like the other in-card section titles.
# The EN/DE toggle st.rerun()s BEFORE this widget renders, which makes Streamlit
# drop widget-keyed state → mirror the choice into a plain session key instead.
st.markdown(f"**{T['filing_type'][lang]}**")
_ft_prev = st.session_state.get("filing_type_persist", "neumeldung")
filing_type = st.radio(
    T["filing_type"][lang],
    ["neumeldung", "korrektur"],
    index=1 if _ft_prev == "korrektur" else 0,
    format_func=lambda x: T["mode_neumeldung"][lang] if x == "neumeldung" else T["mode_korrektur"][lang],
    horizontal=True,
    key="filing_type",
    label_visibility="collapsed",
)
st.session_state["filing_type_persist"] = filing_type
corr_mode = filing_type == "korrektur"
uploaded = st.file_uploader(
    T["upload_label"][lang],
    type=["xlsx", "xlsm"],
    help=T["upload_help"][lang],
)
orig_file = None
if corr_mode:
    # The global dropzone CSS injects "Drop your GIR template here…" into EVERY
    # uploader — scope a correct headline onto this one via its st-key class.
    _xml_drop_txt = ("Drop the originally submitted XML here, or click to browse"
                     if lang == "EN" else
                     "Original eingereichtes Roh-XML hier ablegen oder klicken")
    st.markdown(
        '<style>.stApp [class*="st-key-girxmlup"] '
        '[data-testid="stFileUploaderDropzoneInstructions"] span::after '
        f'{{ content: "{_xml_drop_txt}"; }}</style>',
        unsafe_allow_html=True,
    )
    with st.container(key="girxmlup"):
        orig_file = st.file_uploader(
            T["korr_upload_label"][lang],
            type=["xml"],
            help=T["korr_upload_help"][lang],
        )

# ── Seed Step-2 widget defaults (once), then auto-fill from sheet 1 on upload ──
_WIDGET_DEFAULTS = {
    "f_company":   "PLACEHOLDER_COMPANY_AG",
    "f_tin":       "CHE-123456789",
    "f_tin_iss":   COUNTRY_DISPLAY[_country_idx("CH")],
    "f_jur":       COUNTRY_DISPLAY[_country_idx("CH")],
    "f_ccy":       "CHF",
    "f_fas":       FAS_OPTIONS[0],
    "f_pstart":    date(2024, 1, 1),
    "f_pend":      date(2024, 12, 31),
    "f_rec":       COUNTRY_DISPLAY[_country_idx("CH")],
    "f_role":      "GIR401",
    "f_tintype":   "GIR3001",
    "f_cfs":       "GIR501",
    "f_uperules":  "GIR204",
    "f_upestatus": "GIR301",
}
for _k, _v in _WIDGET_DEFAULTS.items():
    st.session_state.setdefault(_k, _v)

mne_entities: list = st.session_state.get("mne_entities", [])
if uploaded is not None:
    _file_bytes = uploaded.getvalue()
    _sig = (uploaded.name, len(_file_bytes))
    if st.session_state.get("_loaded_sig") != _sig:
        try:
            _wb = openpyxl.load_workbook(io.BytesIO(_file_bytes), data_only=True)
            if MNE_SHEET not in _wb.sheetnames or not _computation_sheets(_wb):
                raise KeyError(f'{MNE_SHEET} / {COMP_SHEET}')
            _mne = read_mne_info(_wb)
            st.session_state["mne_entities"] = _mne["constituent_entities"]
            st.session_state["safe_harbours"] = read_safe_harbours(_wb)
            mne_entities = _mne["constituent_entities"]

            def _opt(val, options, fallback):
                return val if val in options else fallback

            st.session_state["f_company"]   = _mne["company_name"] or "PLACEHOLDER_COMPANY_AG"
            st.session_state["f_tin"]        = _mne["tin_value"] or "CHE-123456789"
            st.session_state["f_tin_iss"]    = COUNTRY_DISPLAY[_country_idx(_mne["tin_issued_by"] or "CH")]
            st.session_state["f_jur"]        = COUNTRY_DISPLAY[_country_idx(_mne["jurisdiction"] or "CH")]
            st.session_state["f_ccy"]        = _opt(_mne["currency"], CURRENCIES, "CHF")
            st.session_state["f_fas"]        = _opt(_mne["fas"], FAS_OPTIONS, FAS_OPTIONS[0])
            if _mne["period_start"]:
                st.session_state["f_pstart"] = date.fromisoformat(_mne["period_start"])
            if _mne["period_end"]:
                st.session_state["f_pend"]   = date.fromisoformat(_mne["period_end"])
            st.session_state["f_role"]       = _opt(_mne["filing_role"], ["GIR401", "GIR402", "GIR404"], "GIR401")
            st.session_state["f_uperules"]   = _opt(_mne["upe_rules"], UPE_RULES_OPTIONS, "GIR204")
            st.session_state["f_upestatus"]  = _opt(_mne["upe_globe_status"], UPE_GLOBE_STATUS_OPTIONS, "GIR301")
            st.session_state["_loaded_sig"]  = _sig
            st.success(T["autofilled"][lang])
        except Exception as e:
            logging.exception("Auto-fill from sheet 1 failed")
            st.warning(T["autofill_failed"][lang].format(e))

_card_close(_card1)

# ── Step 2: Company details ───────────────────────────────────────────────────
# The mode radio is rendered later in this run; read its persisted value (default
# = test, matching the radio's index=1) so the Testmode badge shows immediately.
_testmode = st.session_state.get("submission_mode", "test") == "test"
_card2 = _card_open("girc2")
step_header(2, T["step2"][lang],
            badge=T["testmode_badge"][lang] if _testmode else None)
if mne_entities:
    st.caption(T["ce_count"][lang].format(len(mne_entities)))

col1, col2 = st.columns(2)
with col1:
    company_name  = st.text_input(T["company_name"][lang], key="f_company")
    tin_value     = st.text_input(T["tin"][lang], key="f_tin")
    _tin_sel      = st.selectbox(T["tin_issued_by"][lang], COUNTRY_DISPLAY, key="f_tin_iss")
    tin_issued_by = _tin_sel[:2]
    _jur_sel      = st.selectbox(T["jurisdiction"][lang], COUNTRY_DISPLAY, key="f_jur")
    jurisdiction  = _jur_sel[:2]

with col2:
    currency         = st.selectbox(T["currency"][lang], CURRENCIES, key="f_ccy")
    fas              = st.selectbox(T["fas_label"][lang], FAS_OPTIONS, key="f_fas")
    _period_start_dt = st.date_input(T["period_start"][lang], key="f_pstart")
    _period_end_dt   = st.date_input(T["period_end"][lang],   key="f_pend")
    period_start     = _period_start_dt.strftime("%Y-%m-%d")
    period_end       = _period_end_dt.strftime("%Y-%m-%d")

_rec_sel     = st.selectbox(
    T["partner_country"][lang],
    COUNTRY_DISPLAY,
    key="f_rec",
    help=T["partner_help"][lang],
)
rec_jur_code = _rec_sel[:2]

with st.expander(T["advanced"][lang]):
    reporting_role = st.selectbox(
        T["filing_role"][lang],
        ["GIR401", "GIR402", "GIR404"],
        key="f_role",
        format_func=lambda x: {
            "GIR401": T["gir401"][lang],
            "GIR402": T["gir402"][lang],
            "GIR404": T["gir404"][lang],
        }[x],
        help=T["filing_role_help"][lang],
    )
    tin_type = st.selectbox(
        T["tin_type_label"][lang],
        ["GIR3001", "GIR3002"],
        key="f_tintype",
        format_func=lambda x: {
            "GIR3001": T["gir3001"][lang],
            "GIR3002": T["gir3002"][lang],
        }[x],
        help=T["tin_type_help"][lang],
    )
    cfs_of_upe = st.selectbox(
        T["cfs_upe"][lang],
        ["GIR501", "GIR502", "GIR503"],
        key="f_cfs",
        format_func=lambda x: {
            "GIR501": T["gir501"][lang],
            "GIR502": T["gir502"][lang],
            "GIR503": T["gir503"][lang],
        }[x],
        help=T["cfs_upe_help"][lang],
    )
    upe_rules = st.selectbox(
        T["upe_rules_label"][lang],
        UPE_RULES_OPTIONS,
        key="f_uperules",
        format_func=lambda x: UPE_RULES_LABELS[x],
        help=T["upe_rules_help"][lang],
    )
    upe_globe_status = st.selectbox(
        T["upe_globe_status_label"][lang],
        UPE_GLOBE_STATUS_OPTIONS,
        key="f_upestatus",
        format_func=lambda x: UPE_GLOBE_STATUS_LABELS[x],
        help=T["upe_globe_status_help"][lang],
    )
    submission_mode = st.radio(
        T["submission_mode"][lang],
        ["production", "test"],
        format_func=lambda x: T["mode_production"][lang] if x == "production" else T["mode_test"][lang],
        index=1,
        help=T["mode_help"][lang],
        horizontal=True,
        key="submission_mode",
    )
    if submission_mode == "test":
        st.warning(T["mode_test_banner"][lang])

# ── Constituent-entity TIN editor (fill missing TINs in-app) ──────────────────
# `edited_entities` is what build_xml uses; edits here override the Excel.
edited_entities = mne_entities
if mne_entities:
    st.markdown(f"**{T['ce_table_title'][lang]}**")
    st.caption(T["ce_table_help"][lang])
    _rows = [{
        T["ce_col_entity"][lang]: e["name"],
        T["ce_col_jur"][lang]:    e.get("iso") or "",
        T["ce_col_tin"][lang]:    e.get("tin") or "",
    } for e in mne_entities]
    _edited = st.data_editor(
        _rows,
        disabled=[T["ce_col_entity"][lang], T["ce_col_jur"][lang]],
        hide_index=True,
        use_container_width=True,
        key=f"ce_editor_{st.session_state.get('_loaded_sig', 'none')}",
    )
    edited_entities = []
    for _orig, _row in zip(mne_entities, _edited):
        _e = dict(_orig)
        _tin = str(_row.get(T["ce_col_tin"][lang], "") or "").strip()
        _e["tin"] = _tin or None
        edited_entities.append(_e)
    st.session_state["mne_entities_edited"] = edited_entities
    _n_missing = sum(1 for _e in edited_entities if not _e["tin"])
    if _n_missing:
        st.warning(T["ce_missing_warn"][lang].format(_n_missing))
    else:
        st.caption(T["ce_all_have_tin"][lang])

# ── Safe-harbour jurisdictions — clean read-only summary + advanced subset picker
_SH_CODE_LABELS = {
    "GIR1201": "De minimis exclusion",
    "GIR1202": "QDMTT safe harbour",
    "GIR1203": "Transitional CbCR — De minimis test",
    "GIR1204": "Transitional CbCR — ETR test",
    "GIR1205": "Transitional CbCR — Routine profit test",
    "GIR1206": "Transitional UTPR safe harbour",
    "GIR1207": "Permanent SH — De minimis test",
    "GIR1208": "Permanent SH — ETR test",
    "GIR1209": "Permanent SH — Routine profit test",
}
_safe_harbours = st.session_state.get("safe_harbours", [])
if _safe_harbours:
    st.markdown(f"**{T['sh_title'][lang]}**")
    st.caption(T["sh_help"][lang].format(len(_safe_harbours)))
    _excluded = st.session_state.setdefault("sh_excluded", set())
    # Make the row ✕ a small, borderless glyph instead of a boxed button.
    st.markdown("""
    <style>
    div[data-testid="stButton"] button[kind="tertiary"] {
        color:#b9c2cb; padding:0 4px; min-height:0; line-height:1; font-size:0.9rem;
    }
    div[data-testid="stButton"] button[kind="tertiary"]:hover { color:#e0303c; }
    </style>""", unsafe_allow_html=True)
    # Header row + one row per included jurisdiction, each with a ✕ to remove it.
    _h1, _h2, _h3 = st.columns([6, 7, 1])
    _h1.markdown(f"<span style='color:#6a7681;font-size:0.8rem;'>{T['ce_col_jur'][lang]}</span>",
                 unsafe_allow_html=True)
    _h2.markdown("<span style='color:#6a7681;font-size:0.8rem;'>"
                 + ("Safe harbour" if lang == "EN" else "Safe Harbour") + "</span>", unsafe_allow_html=True)
    for _i, sh in enumerate(_safe_harbours):
        _iso = sh.get("iso") or "?"
        if _iso in _excluded:
            continue
        _c1, _c2, _c3 = st.columns([6, 7, 1], vertical_alignment="center")
        _c1.write(f"{_iso} — {sh.get('name', '')}")
        _c2.write(f"{sh['sh_code']} — {_SH_CODE_LABELS.get(sh['sh_code'], '')}")
        if _c3.button("✕", key=f"rm_{_i}_{_iso}", type="tertiary",
                      help="Remove this jurisdiction from the file" if lang == "EN"
                           else "Diese Jurisdiktion aus der Datei entfernen"):
            _excluded.add(_iso)
            st.rerun()
    if _excluded:
        _r1, _r2 = st.columns([5, 2])
        _r1.caption((f"{len(_excluded)} removed: " if lang == "EN" else f"{len(_excluded)} entfernt: ")
                    + ", ".join(sorted(_excluded)))
        if _r2.button("↩ Reset" if lang == "EN" else "↩ Zurücksetzen", key="sh_reset"):
            _excluded.clear()
            st.rerun()
    st.session_state["sh_selected_isos"] = [
        (sh.get("iso") or "?") for sh in _safe_harbours if (sh.get("iso") or "?") not in _excluded
    ]

_card_close(_card2)

# ── Step 3: Export ────────────────────────────────────────────────────────────
_card3 = _card_open("girc3")
step_header(3, T["step3"][lang])
st.write("")

def validate_inputs(cfg: dict) -> list[str]:
    L = st.session_state.get("lang", "EN")
    errors = []
    if not re.match(r"^[A-Z]{2}$", cfg["jurisdiction"]):
        errors.append(T["err_jurisdiction"][L])
    if not re.match(r"^[A-Z]{3}$", cfg["currency"]):
        errors.append(T["err_currency"][L])
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", cfg["period_start"]):
        errors.append(T["err_period_start"][L])
    if not re.match(r"^\d{4}-\d{2}-\d{2}$", cfg["period_end"]):
        errors.append(T["err_period_end"][L])
    if not cfg["company_name"].strip():
        errors.append(T["err_company"][L])
    if not cfg["tin_value"].strip():
        errors.append(T["err_tin"][L])
    if not re.match(r"^[A-Z]{2}$", cfg["tin_issued_by"]):
        errors.append(T["err_tin_issued"][L])
    if not re.match(r"^[A-Z]{2}$", cfg["rec_jur_code"]):
        errors.append(T["err_rec_jur"][L])
    return errors


def _current_cfg(entities) -> dict:
    """Assemble the build_xml cfg from the Step-2 widget state (correction mode)."""
    return {
        "company_name":    company_name,
        "tin_value":       tin_value,
        "tin_issued_by":   tin_issued_by,
        "tin_type":        tin_type,
        "reporting_role":  reporting_role,
        "rec_jur_code":    rec_jur_code.strip().upper(),
        "currency":        currency,
        "jurisdiction":    jurisdiction,
        "fas":             fas,
        "cfs_of_upe":      cfs_of_upe,
        "period_start":    period_start,
        "period_end":      period_end,
        "upe_rules":       upe_rules,
        "upe_globe_status": upe_globe_status,
        "constituent_entities": entities,
    }


@st.cache_data(show_spinner=False)
def _corr_regenerate(file_bytes: bytes, cfg_json: str, sel_isos_json: str, test: bool) -> str:
    """Regenerate the full report from the corrected Excel (cached — reruns on
    every widget change would otherwise re-read the workbook each time)."""
    parsed = read_excel_v2(file_bytes)
    cfg = json.loads(cfg_json)
    _sel = json.loads(sel_isos_json)
    cfg["safe_harbours"] = parsed["safe_harbours"] if _sel is None else [
        s for s in parsed["safe_harbours"] if (s.get("iso") or "?") in _sel]
    if not cfg.get("constituent_entities"):
        cfg["constituent_entities"] = parsed["mne"]["constituent_entities"]
    return build_xml(parsed["computations"], cfg, test_mode=test)


# ── Correction mode: parse original + live section diff + selection ──────────
_corr_ctx = None
_corr_rows = []
_corr_selection: dict = {}
_corr_new_xml = None
if corr_mode and uploaded is not None and orig_file is not None:
    try:
        _corr_ctx = parse_original_xml(orig_file.getvalue())
        st.caption(T["korr_parse_ok"][lang].format(
            len(_corr_ctx["sections"]) + 1, _corr_ctx["message_ref_id"]))
        _corr_new_xml = _corr_regenerate(
            uploaded.getvalue(),
            json.dumps(_current_cfg(edited_entities), sort_keys=True),
            json.dumps(st.session_state.get("sh_selected_isos")),
            submission_mode == "test",
        )
        _corr_rows = diff_correction_sections(_corr_ctx, _corr_new_xml)

        st.markdown(f"**{T['korr_sections_title'][lang]}**")
        st.caption(T["korr_sections_help"][lang])
        _n_new = sum(1 for r in _corr_rows if r["status"] == "new")
        if _n_new:
            st.warning(T["korr_new_blocked"][lang].format(_n_new))
        _ST_COLOR = {"changed": "#e0303c", "unchanged": "#1f9d57",
                     "new": "#9a6700", "missing": "#9a6700"}
        for r in _corr_rows:
            _c1, _c2, _c3 = st.columns([5, 4, 3], vertical_alignment="center")
            _lbl = r["type"] + (f" — {r['jurisdiction']}" if r["jurisdiction"] else "")
            if r["status"] in ("changed", "unchanged"):
                if _c1.checkbox(_lbl, value=(r["status"] == "changed"),
                                key=f"korrsel_{r['key']}"):
                    _corr_selection[r["key"]] = "correct"
            else:
                _c1.markdown(f"<span style='color:#8a96a3;'>{_lbl}</span>",
                             unsafe_allow_html=True)
            _c2.markdown(f"<span style='color:{_ST_COLOR[r['status']]};font-size:0.82rem;"
                         f"font-weight:600;'>{T['korr_st_' + r['status']][lang]}</span>",
                         unsafe_allow_html=True)
            _c3.markdown(f"<span style='color:#8a96a3;font-size:0.72rem;'>"
                         f"{(r['doc_ref_id'] or '')[:13]}…</span>", unsafe_allow_html=True)

        with st.expander(T["korr_advanced"][lang]):
            st.caption(T["korr_adv_help"][lang])
            _fi_dr = st.text_input(T["korr_fi_docref"][lang],
                                   value=_corr_ctx["filing_doc_ref_id"], key="korr_fi_dr")
            if _fi_dr.strip():
                _corr_ctx["filing_doc_ref_id"] = _fi_dr.strip()
            for _s in _corr_ctx["sections"]:
                _slbl = _s["type"] + (f" — {_s['jurisdiction']}" if _s["jurisdiction"] else "")
                _dr = st.text_input(f"DocRefId: {_slbl}", value=_s["doc_ref_id"],
                                    key=f"korr_dr_{_s['key']}")
                if _dr.strip():
                    _s["doc_ref_id"] = _dr.strip()
            for _r in _corr_rows:
                if _r["status"] not in ("missing", "unchanged"):
                    continue
                _slbl = _r["type"] + (f" — {_r['jurisdiction']}" if _r["jurisdiction"] else "")
                if st.checkbox(T["korr_storno_label"][lang].format(_slbl),
                               key=f"korrdel_{_r['key']}"):
                    _corr_selection[_r["key"]] = "delete"
    except Exception as _e:
        logging.exception("Correction pre-processing failed")
        st.error(T["korr_parse_err"][lang].format(_e))
        _corr_ctx = None
elif corr_mode:
    st.info(T["korr_need_both"][lang])

_gen_disabled = uploaded is None or (corr_mode and (orig_file is None or _corr_ctx is None))
_gen_clicked = st.button(T["generate_btn"][lang], type="primary", disabled=_gen_disabled)

if _gen_clicked and corr_mode:
    if _corr_ctx is None or _corr_new_xml is None:
        st.error(T["korr_need_both"][lang])
    elif not _corr_selection:
        st.error(T["korr_none_selected"][lang])
    else:
        with st.spinner(T["spinner_gen"][lang]):
            try:
                input_errors = validate_inputs(_current_cfg(edited_entities))
                if input_errors:
                    for err in input_errors:
                        st.error(err)
                    st.stop()
                xml_str = apply_correction(_corr_new_xml, _corr_ctx, _corr_selection,
                                           test_mode=(submission_mode == "test"))
                _n_corr = sum(1 for v in _corr_selection.values() if v == "correct")
                _n_del  = sum(1 for v in _corr_selection.values() if v == "delete")
                _prefix = "Test_" if submission_mode == "test" else ""
                st.session_state["xml_str"] = xml_str
                st.session_state["xml_filename"] = (
                    f"{_prefix}gir_{period_end[:4]}_{jurisdiction}_korrektur.xml")
                st.caption(T["korr_summary"][lang].format(_n_corr, _n_del))

                l1, l2 = validate_correction(xml_str, _corr_ctx)
                all_ok = all(ok for _, ok, _ in l1) and all(ok for _, ok, _ in l2)
                st.session_state["validation_ok"] = all_ok

                for _title, _checks in ((T["korr_l1_title"][lang], l1),
                                        (T["korr_l2_title"][lang], l2)):
                    _np, _nt = sum(1 for _, ok, _ in _checks if ok), len(_checks)
                    with st.expander(
                        f"{'✅' if _np == _nt else '⚠️'}  {_title} — {_np}/{_nt} {T['checks_passed'][lang]}",
                        expanded=_np != _nt,
                    ):
                        for label, ok, detail in _checks:
                            icon = "✅" if ok else "❌"
                            if detail and not ok:
                                st.markdown(f"{icon} &nbsp; **{label}**  \n"
                                            f"&nbsp;&nbsp;&nbsp;&nbsp;`{detail}`")
                            else:
                                st.markdown(f"{icon} &nbsp; {label}")

                if all_ok:
                    st.success(T["all_passed"][lang])
                else:
                    st.warning(T["download_blocked"][lang])
                with st.expander(T["preview_xml"][lang]):
                    st.code(xml_str, language="xml")
            except Exception as e:
                logging.exception("Correction build failed")
                st.error(T["error_msg"][lang].format(e))
elif _gen_clicked:
    if uploaded is None:
        st.error(T["upload_first"][lang])
    else:
        with st.spinner(T["spinner_gen"][lang]):
            try:
                file_bytes = uploaded.getvalue()
                parsed = read_excel_v2(file_bytes)
                computations = parsed["computations"]
                data   = parsed["data"]      # first computation (single-jur back-compat)
                # Use the (possibly TIN-edited) entities from the in-app table,
                # falling back to the freshly parsed list.
                ce_list = edited_entities or parsed["mne"]["constituent_entities"]
                # Emit only the safe-harbour jurisdictions selected in the UI
                # (default = all). Lets us submit minimal files to bisect ESTV issues.
                _sel = st.session_state.get("sh_selected_isos")
                sh_list = parsed["safe_harbours"] if _sel is None else [
                    s for s in parsed["safe_harbours"] if (s.get("iso") or "?") in _sel
                ]

                cfg = {
                    "company_name":    company_name,
                    "tin_value":       tin_value,
                    "tin_issued_by":   tin_issued_by,
                    "tin_type":        tin_type,
                    "reporting_role":  reporting_role,
                    "rec_jur_code":    rec_jur_code.strip().upper(),
                    "currency":        currency,
                    "jurisdiction":    jurisdiction,
                    "fas":             fas,
                    "cfs_of_upe":      cfs_of_upe,
                    "period_start":    period_start,
                    "period_end":      period_end,
                    "upe_rules":       upe_rules,
                    "upe_globe_status": upe_globe_status,
                    "constituent_entities": ce_list,
                    "safe_harbours":        sh_list,
                }

                input_errors = validate_inputs(cfg)
                if input_errors:
                    for err in input_errors:
                        st.error(err)
                    st.stop()

                xml_str = build_xml(computations, cfg, test_mode=(submission_mode == "test"))
                st.session_state["xml_str"]      = xml_str
                st.session_state["xml_filename"] = f"gir_{period_end[:4]}_{jurisdiction}.xml"

                for _d in computations:
                    if len(computations) > 1:
                        st.markdown(f"**{_d.get('jur_iso') or jurisdiction}**")
                    cols = st.columns(4)
                    cols[0].metric("AdjustedFANIL",  f"{_d['adjusted_fanil']:,}")
                    cols[1].metric("NetGlobeIncome",  f"{_d['net_globe_income']:,}")
                    cols[2].metric("AdjustedCovTax",  f"{_d['adjusted_cov_tax']:,}")
                    cols[3].metric("ETR",             fmt_etr(_d["adjusted_cov_tax"], _d["net_globe_income"]))

                checks  = validate_xml(xml_str)
                n_pass  = sum(1 for _, ok, _ in checks if ok)
                n_total = len(checks)
                all_ok  = n_pass == n_total
                st.session_state["validation_ok"] = all_ok

                with st.expander(
                    f"{'✅' if all_ok else '⚠️'}  {T['validation_title'][lang]} — "
                    f"{n_pass}/{n_total} {T['checks_passed'][lang]}",
                    expanded=not all_ok,
                ):
                    for label, ok, detail in checks:
                        icon = "✅" if ok else "❌"
                        if detail and not ok:
                            st.markdown(f"{icon} &nbsp; **{label}**  \n"
                                        f"&nbsp;&nbsp;&nbsp;&nbsp;`{detail}`")
                        else:
                            st.markdown(f"{icon} &nbsp; {label}")
                    if not all_ok:
                        st.caption(T["fix_issues"][lang])

                if all_ok:
                    st.success(T["all_passed"][lang])
                else:
                    st.warning(T["download_blocked"][lang])

                # Plain language summary
                with st.expander(T["summary_label"][lang]):
                    C = cfg["currency"]

                    def _row(label, value, bold=False):
                        ca, cb = st.columns([2, 3])
                        ca.markdown(
                            f"<span style='color:#6a7681;font-size:0.84rem;'>{label}</span>",
                            unsafe_allow_html=True,
                        )
                        color = "#e0303c" if bold else "#313c45"
                        weight = "700" if bold else "600"
                        cb.markdown(
                            f"<span style='color:{color};font-size:0.84rem;font-weight:{weight};'>{value}</span>",
                            unsafe_allow_html=True,
                        )

                    st.markdown(f"**{T['sum_filing'][lang]}**")
                    _row(T["sum_company"][lang],     cfg["company_name"])
                    _row(T["sum_tin"][lang],         cfg["tin_value"])
                    _row(T["sum_jurisdiction"][lang],cfg["jurisdiction"])
                    _row(T["sum_period"][lang],      f"{cfg['period_start']} – {cfg['period_end']}")
                    _row(T["sum_currency"][lang],    cfg["currency"])
                    _row(T["sum_fas"][lang],         cfg["fas"])
                    _row(T["sum_role"][lang],        cfg["reporting_role"])
                    _row(T["sum_partner"][lang],     cfg["rec_jur_code"])

                    for data in computations:
                        if len(computations) > 1:
                            st.markdown(
                                f"<div style='margin-top:8px;font-weight:700;color:#e0303c;"
                                f"font-size:0.9rem;'>{data.get('jur_iso') or cfg['jurisdiction']}</div>",
                                unsafe_allow_html=True,
                            )
                        st.markdown("---")
                        st.markdown(f"**{T['sum_income'][lang]}**")
                        _row(T["sum_fanil"][lang], f"{data['adjusted_fanil']:,} {C}")

                        nz_income = [(code, amt) for code, amt in data["income_adj"] if amt != 0]
                        if nz_income:
                            st.caption(T["sum_income_adj"][lang])
                            for code, amt in nz_income:
                                desc = GIR_INCOME_LABELS.get(code, code)
                                ca, cb = st.columns([2, 3])
                                ca.markdown(
                                    f"<span style='color:#6a7681;font-size:0.78rem;'>{code} — {desc}</span>",
                                    unsafe_allow_html=True,
                                )
                                cb.markdown(
                                    f"<span style='font-size:0.78rem;'>{amt:,} {C}</span>",
                                    unsafe_allow_html=True,
                                )

                        _row(T["sum_net_income"][lang], f"{data['net_globe_income']:,} {C}", bold=True)

                        st.markdown("---")
                        st.markdown(f"**{T['sum_tax'][lang]}**")
                        _row(T["sum_curr_tax"][lang], f"{data['aggregate_curr_tax']:,} {C}")

                        nz_tax = [(code, amt) for code, amt in data["cov_tax_adj"] if amt != 0]
                        if nz_tax:
                            st.caption(T["sum_tax_adj"][lang])
                            for code, amt in nz_tax:
                                desc = GIR_TAX_LABELS.get(code, code)
                                ca, cb = st.columns([2, 3])
                                ca.markdown(
                                    f"<span style='color:#6a7681;font-size:0.78rem;'>{code} — {desc}</span>",
                                    unsafe_allow_html=True,
                                )
                                cb.markdown(
                                    f"<span style='font-size:0.78rem;'>{amt:,} {C}</span>",
                                    unsafe_allow_html=True,
                                )

                        _row(T["sum_adj_tax"][lang], f"{data['adjusted_cov_tax']:,} {C}", bold=True)

                        st.markdown("---")
                        st.markdown(f"**{T['sum_result'][lang]}**")
                        _etr    = fmt_etr(data["adjusted_cov_tax"], data["net_globe_income"])
                        etr_pct = float(_etr) * 100
                        _row(T["sum_etr"][lang], f"{_etr} ({etr_pct:.2f}%)", bold=True)

                with st.expander(T["preview_xml"][lang]):
                    st.code(xml_str, language="xml")

            except KeyError as e:
                logging.exception("Sheet not found during Excel read")
                st.error(T["sheet_not_found"][lang].format(e))
            except Exception as e:
                logging.exception("XML generation failed")
                st.error(T["error_msg"][lang].format(e))

elif uploaded is None and not corr_mode:
    st.info(T["upload_to_enable"][lang])

_card_close(_card3)

# ── Step 4: Encrypt for ESTV ──────────────────────────────────────────────────
_card4 = _card_open("girc4")
step_header(4, T["step4"][lang])

xml_ready = "xml_str" in st.session_state
validation_ok = st.session_state.get("validation_ok", False)

# Determine which PEM to use
pem_override = None

st.markdown(
    f"<div style='background:#f5f9fc; border-left:3px solid #e0303c; padding:10px 14px;"
    f"border-radius:4px; margin-bottom:12px; font-size:0.88rem; color:#313c45;'>"
    f"🔑 &nbsp;{T['bundled_key_info'][lang]}</div>",
    unsafe_allow_html=True,
)

with st.expander(T["override_key"][lang]):
    pem_file = st.file_uploader(T["pem_label"][lang], type=["pem", "cer"])
    if pem_file is not None:
        pem_override = pem_file.read()
        st.success(T["override_active"][lang])

pem_bytes_to_use = pem_override if pem_override else _BUNDLED_PEM

_enc_col, _raw_col = st.columns([1, 1])
with _enc_col:
    _do_encrypt = st.button(
        T["encrypt_btn"][lang],
        type="primary",
        disabled=(not xml_ready or not validation_ok or pem_bytes_to_use is None),
    )
with _raw_col:
    st.download_button(
        label=T["download_xml"][lang],
        data=(st.session_state.get("xml_str", "") or "").encode("utf-8"),
        file_name=st.session_state.get("xml_filename", "gir.xml"),
        mime="application/xml",
        disabled=(not xml_ready or not validation_ok),
    )

if _do_encrypt:
    if not xml_ready:
        st.error(T["generate_first"][lang])
    elif pem_bytes_to_use is None:
        st.error(T["upload_pem"][lang])
    else:
        with st.spinner(T["spinner_enc"][lang]):
            try:
                zip_bytes    = encrypt_for_estv(st.session_state["xml_str"], pem_bytes_to_use)
                base_name    = st.session_state["xml_filename"].replace(".xml", "")
                zip_filename = f"{base_name}_encrypted.zip"
                st.download_button(
                    label=T["download_zip"][lang],
                    data=zip_bytes,
                    file_name=zip_filename,
                    mime="application/zip",
                )
                st.success(T["ready_estv"][lang])
            except Exception as e:
                logging.exception("Encryption failed")
                st.error(T["enc_failed"][lang].format(e))

if not xml_ready:
    st.info(T["gen_first_long"][lang])
elif not validation_ok:
    st.warning(T["encrypt_blocked"][lang])

_card_close(_card4)

st.divider()
with st.expander(T["disclaimer_label"][lang]):
    st.markdown(
        f"<p style='color:#6a7681; font-size:0.85rem;'>{T['disclaimer_text'][lang]}</p>",
        unsafe_allow_html=True,
    )
st.markdown(
    "<p style='color:#6a7681; font-size:0.75rem; margin:0;'>MME Legal | Tax | Compliance</p>",
    unsafe_allow_html=True,
)
